from datetime import datetime
from io import BytesIO
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
PROGRAM_COLUMN = "Programs Name"
UID_COLUMN = "Clients Unique Identifier"
DATE_COLUMN = "Services Start Date Date"
USER_COLUMN = "Services User Creating"
NOTES_COLUMN = "Services Services Notes"
SERVICE_NAME_COLUMN = "Services Name"
ITEM_COLUMNS = ["Service Items Service Item Name", "Services Service Item Name"]
REQUIRED_COLUMNS = [
    PROGRAM_COLUMN,
    UID_COLUMN,
    "Clients First Name",
    "Clients Last Name",
    "Services Service Added Date",
    DATE_COLUMN,
    "Services Service Reporting Period Start Date",
    "Services Service Reporting Period End Date",
    "Services Service Category",
    "Services Name",
    USER_COLUMN,
    "Services Type Provided",
    NOTES_COLUMN,
]


class ServicesService:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def process(
        self,
        program,
        uploaded_file,
        start_date,
        end_date,
        structure,
        current_metrics,
        service_users=None,
        service_user_teams=None,
    ):
        program = str(program).strip().upper()
        if program not in {"MHRT", "RRT"}:
            raise ValueError("Program must be MHRT or RRT.")
        start = self._parse_date(start_date, "start")
        end = self._parse_date(end_date, "end")
        if start > end:
            raise ValueError("Start date must be on or before end date.")

        filename = Path(uploaded_file.filename).name
        self._validate_filename_program(program, filename)
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError("Services upload must be a CSV, XLSX, or XLS file.")
        content = uploaded_file.read()
        if not content:
            raise ValueError("The uploaded Services file is empty.")

        source = self._load_services_frame(content, extension)
        filtered, item_col = self._prepare_frame(source, program, start, end)
        rows_before_dedupe = len(filtered)
        dedupe_columns = [UID_COLUMN, USER_COLUMN, NOTES_COLUMN, DATE_COLUMN, item_col]
        filtered = filtered.drop_duplicates(subset=dedupe_columns, keep="first").copy()

        staff_lookup, team_lookup = self._structure_lookups(structure)
        staff_to_service_user = service_users or {}
        service_user_lookup = {
            str(username).strip().casefold(): staff
            for staff, username in staff_to_service_user.items()
            if str(staff).strip() and str(username).strip()
        }

        def resolve_staff(service_user):
            key = service_user.casefold()
            return service_user_lookup.get(key, staff_lookup.get(key, service_user))

        filtered["Fullname"] = filtered[USER_COLUMN].map(resolve_staff)
        mapped_teams = {
            str(staff).strip().casefold(): str(team).strip()
            for staff, team in (service_user_teams or {}).items()
            if str(staff).strip() and str(team).strip()
        }
        manager_names = {
            self._team_key(manager.get("name", "")): str(manager.get("name", "")).strip()
            for manager in structure.get("managers", [])
        }

        def resolve_team(staff):
            current_team = team_lookup.get(staff)
            if current_team:
                return current_team
            desired = mapped_teams.get(staff.casefold())
            if desired:
                return manager_names.get(self._team_key(desired), desired)
            return team_lookup.get(staff, "No Team") or "No Team"

        filtered["Team"] = filtered["Fullname"].map(resolve_team)
        unmatched_service_users = sorted(
            {
                service_user
                for service_user in filtered[USER_COLUMN].tolist()
                if service_user.casefold() not in service_user_lookup
                and service_user.casefold() not in staff_lookup
            },
            key=str.casefold,
        )
        displayed_service_users = dict(staff_to_service_user)
        for service_user in unmatched_service_users:
            displayed_service_users.setdefault(service_user, service_user)

        total_summary = filtered.groupby("Fullname", dropna=False).size().to_dict()
        shelter_mask = filtered[item_col].str.casefold() == "shelter referral"
        shelter_frame = filtered[shelter_mask].copy()
        shelter_summary = shelter_frame.groupby("Fullname", dropna=False).size().to_dict()

        report_staff = list(dict.fromkeys(filtered["Fullname"].tolist()))
        report_team_lookup = (
            filtered.drop_duplicates(subset=["Fullname"], keep="first")
            .set_index("Fullname")["Team"]
            .to_dict()
        )
        roster_staff = {
            staff.casefold()
            for manager in structure.get("managers", [])
            for staff in manager.get("staff", [])
        }
        unmatched_staff = sorted(
            {staff for staff in report_staff if staff.casefold() not in roster_staff},
            key=str.casefold,
        )
        metrics = {name: dict(values) for name, values in current_metrics.items()}
        all_staff = [staff for manager in structure.get("managers", []) for staff in manager.get("staff", [])]
        metric_staff = list(dict.fromkeys(all_staff + report_staff))
        for staff in metric_staff:
            metrics.setdefault(staff, {})
            metrics[staff]["services"] = int(total_summary.get(staff, 0))
            metrics[staff]["referrals_to_shelter"] = int(shelter_summary.get(staff, 0))

        output_name = self._write_workbook(program, filtered, shelter_frame, item_col, start, end)
        return {
            "structure": structure,
            "metrics": metrics,
            "download_name": output_name,
            "source_filename": filename,
            "rows_in_date_range": int(len(filtered)),
            "shelter_referrals": int(len(shelter_frame)),
            "duplicates_removed": int(rows_before_dedupe - len(filtered)),
            "staff_count": int(filtered["Fullname"].nunique()),
            "report_staff": report_staff,
            "unmatched_staff": unmatched_staff,
            "unmatched_service_users": unmatched_service_users,
            "service_users": displayed_service_users,
        }

    def _load_services_frame(self, content, extension):
        stream = BytesIO(content)
        if extension == ".csv":
            try:
                return pd.read_csv(stream)
            except UnicodeDecodeError:
                stream.seek(0)
                return pd.read_csv(stream, encoding="latin-1")

        workbook = pd.ExcelFile(stream)
        for sheet_name in workbook.sheet_names:
            preview = pd.read_excel(workbook, sheet_name=sheet_name, header=None, nrows=30)
            for header_row in range(len(preview)):
                values = {
                    self._clean_text(value)
                    for value in preview.iloc[header_row].tolist()
                    if pd.notna(value)
                }
                if {UID_COLUMN, USER_COLUMN, DATE_COLUMN}.issubset(values):
                    candidate = pd.read_excel(workbook, sheet_name=sheet_name, header=header_row)
                    candidate.columns = [self._clean_text(column) for column in candidate.columns]
                    return candidate
        raise ValueError("No Excel sheet contains the expected Services columns.")

    def _prepare_frame(self, frame, program, start, end):
        frame = frame.copy()
        frame.columns = [str(column).strip() for column in frame.columns]
        item_col = next((column for column in ITEM_COLUMNS if column in frame.columns), None)
        missing = [column for column in REQUIRED_COLUMNS if column not in frame.columns]
        if not item_col:
            missing.append(ITEM_COLUMNS[0])
        if missing:
            raise ValueError("The Services report is missing columns: " + ", ".join(missing))

        frame[DATE_COLUMN] = pd.to_datetime(frame[DATE_COLUMN], errors="coerce")
        frame = frame[frame[DATE_COLUMN].between(start, end, inclusive="both")].copy()
        program_mask = frame[PROGRAM_COLUMN].map(lambda value: self._matches_program(value, program))
        if frame[PROGRAM_COLUMN].fillna("").astype(str).str.strip().ne("").any():
            rows_in_date_range = len(frame)
            frame = frame[program_mask].copy()
            if rows_in_date_range and frame.empty:
                raise ValueError(f"The Services report has no {program} rows in the selected date range.")

        for column in [UID_COLUMN, USER_COLUMN, NOTES_COLUMN, item_col, SERVICE_NAME_COLUMN]:
            frame[column] = frame[column].fillna("").astype(str).map(self._clean_text)
        frame = frame[
            (frame[UID_COLUMN] != "")
            & (frame[USER_COLUMN] != "")
            & (frame[SERVICE_NAME_COLUMN] != "")
        ].copy()
        return frame, item_col

    @staticmethod
    def _matches_program(value, program):
        text = re.sub(r"[^a-z0-9]+", " ", str(value).casefold()).strip()
        tokens = set(text.split())
        if program == "MHRT":
            return "mhrt" in tokens or "mobile homelessness resolution" in text
        return "rrt" in tokens or "rapid response team" in text

    @staticmethod
    def _validate_filename_program(program, filename):
        normalized = re.sub(r"[^a-z0-9]+", " ", filename.casefold()).strip()
        tokens = set(normalized.split())
        mh_rt_file = "mhrt" in tokens or "mobile homelessness resolution" in normalized
        rrt_file = "rrt" in tokens or "rapid response team" in normalized
        if program == "MHRT" and rrt_file and not mh_rt_file:
            raise ValueError("This appears to be an RRT Services report. Switch the app to RRT before processing it.")
        if program == "RRT" and mh_rt_file and not rrt_file:
            raise ValueError("This appears to be an MHRT Services report. Switch the app to MHRT before processing it.")

    @staticmethod
    def _parse_date(value, label):
        try:
            return pd.Timestamp(value).normalize()
        except (TypeError, ValueError):
            raise ValueError(f"Select a valid {label} date.") from None

    @staticmethod
    def _clean_text(value):
        return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()

    @staticmethod
    def _structure_lookups(structure):
        staff_lookup = {}
        team_lookup = {}
        for manager in structure.get("managers", []):
            manager_name = str(manager.get("name", "")).strip()
            for staff in manager.get("staff", []):
                staff_name = str(staff).strip()
                staff_lookup[staff_name.casefold()] = staff_name
                team_lookup[staff_name] = manager_name
        return staff_lookup, team_lookup

    @staticmethod
    def _add_missing_staff(structure, report_staff, team_lookup):
        result = {
            "program": structure.get("program", ""),
            "managers": [
                {"name": manager.get("name", ""), "staff": list(manager.get("staff", []))}
                for manager in structure.get("managers", [])
            ],
        }
        known = {staff.casefold() for manager in result["managers"] for staff in manager["staff"]}
        no_team = next((manager for manager in result["managers"] if manager["name"].casefold() == "no team"), None)
        for staff in report_staff:
            if staff.casefold() in known:
                continue
            manager_name = team_lookup.get(staff, "No Team") or "No Team"
            manager = next(
                (item for item in result["managers"] if item["name"].casefold() == manager_name.casefold()),
                None,
            )
            if manager is None:
                if manager_name == "No Team" and no_team is not None:
                    manager = no_team
                else:
                    manager = {"name": manager_name, "staff": []}
                    result["managers"].append(manager)
                    if manager_name == "No Team":
                        no_team = manager
            manager["staff"].append(staff)
            known.add(staff.casefold())
        return result

    @staticmethod
    def _team_key(value):
        return re.sub(r"[^a-z0-9]+", "", str(value).casefold())

    @classmethod
    def _sync_report_staff_teams(cls, structure, report_staff, team_lookup):
        result = {
            "program": structure.get("program", ""),
            "active": dict(structure.get("active", {})),
            "managers": [
                {"name": str(manager.get("name", "")).strip(), "staff": list(manager.get("staff", []))}
                for manager in structure.get("managers", [])
            ],
        }
        report_by_key = {staff.casefold(): staff for staff in report_staff}
        for manager in result["managers"]:
            manager["staff"] = [
                staff for staff in manager["staff"]
                if str(staff).strip().casefold() not in report_by_key
            ]

        managers_by_key = {cls._team_key(manager["name"]): manager for manager in result["managers"]}
        for staff in report_staff:
            desired = team_lookup.get(staff, "No Team") or "No Team"
            team_key = cls._team_key(desired)
            manager = managers_by_key.get(team_key)
            if manager is None:
                manager = {"name": desired, "staff": []}
                result["managers"].append(manager)
                managers_by_key[team_key] = manager
            manager["staff"].append(staff)
            result["active"].setdefault(staff, True)
        return result

    def _write_workbook(self, program, filtered, shelter_frame, item_col, start, end):
        program_dir = self.output_dir / program
        program_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{program}_Services_Pivots_{start:%Y%m%d}_{end:%Y%m%d}_{timestamp}.xlsx"
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        self._write_summary_sheet(summary_sheet, filtered, shelter_frame, item_col)
        self._write_data_sheet(workbook, "Filtered", filtered)
        workbook.save(program_dir / filename)
        return filename

    def _write_summary_sheet(self, sheet, filtered, shelter_frame, item_col):
        row = 1
        row = self._write_grouped_pivot(sheet, row, "Total Services Provided", filtered)
        row += 2
        sheet.cell(row, 1, item_col)
        sheet.cell(row, 2, "Shelter Referral")
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor="B7E1F7")
        row += 2
        self._write_grouped_pivot(sheet, row, "Referral to Shelter", shelter_frame)
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 34
        sheet.freeze_panes = "A3"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def _write_grouped_pivot(sheet, start_row, title, frame):
        sheet.cell(start_row, 1, title)
        sheet.cell(start_row, 1).font = Font(bold=True)
        header_row = start_row + 2
        sheet.cell(header_row, 1, "Row Labels")
        sheet.cell(header_row, 2, f"Count of {UID_COLUMN}")
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B7E1F7")
            cell.border = Border(bottom=Side(style="thin", color="5B9BD5"))

        grouped = (
            frame.groupby(["Team", "Fullname"], dropna=False)
            .size()
            .reset_index(name="count")
        )
        grouped["_team_sort"] = grouped["Team"].map(lambda value: (value == "No Team", value.casefold()))
        grouped["_staff_sort"] = grouped["Fullname"].str.casefold()
        grouped = grouped.sort_values(["_team_sort", "_staff_sort"], kind="stable")

        row = header_row + 1
        for team, team_rows in grouped.groupby("Team", sort=False):
            sheet.cell(row, 1, team)
            sheet.cell(row, 2, int(team_rows["count"].sum()))
            for cell in sheet[row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.border = Border(bottom=Side(style="thin", color="5B9BD5"))
            row += 1
            for record in team_rows.to_dict("records"):
                sheet.cell(row, 1, record["Fullname"])
                sheet.cell(row, 1).alignment = Alignment(indent=2)
                sheet.cell(row, 2, int(record["count"]))
                row += 1
        sheet.cell(row, 1, "Grand Total")
        sheet.cell(row, 2, int(len(frame)))
        for cell in sheet[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B7E1F7")
            cell.border = Border(top=Side(style="thin", color="5B9BD5"))
        return row

    @staticmethod
    def _write_data_sheet(workbook, sheet_name, frame):
        sheet = workbook.create_sheet(sheet_name)
        columns = list(frame.columns)
        for column_number, header in enumerate(columns, start=1):
            cell = sheet.cell(1, column_number, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row_number, values in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_number, value in enumerate(values, start=1):
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                cell = sheet.cell(row_number, column_number, value)
                if isinstance(value, datetime):
                    cell.number_format = "yyyy-mm-dd"
        sheet.freeze_panes = "A2"
        last_column = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A1:{last_column}{max(1, len(frame) + 1)}"
        if len(frame):
            table = Table(displayName="FilteredServicesData", ref=f"A1:{last_column}{len(frame) + 1}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)
        for column_number, header in enumerate(columns, start=1):
            values = [str(header)] + [str(value) for value in frame.iloc[:, column_number - 1].dropna().head(200)]
            sheet.column_dimensions[get_column_letter(column_number)].width = min(max(map(len, values)) + 2, 42)
