from datetime import datetime
from io import BytesIO
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
UID_COLUMN = "Clients Unique Identifier"
LAST_NAME_COLUMN = "Clients Last Name"
PROGRAM_COLUMN = "Programs Name"
START_COLUMN = "Enrollments Project Start Date"
EXIT_COLUMN = "Enrollments Project Exit Date"
NOTE_DATE_COLUMN = "Client Notes - Enrollment Level Case Note Date"
ADDED_DATE_COLUMN = "Client Notes - Enrollment Level Date Added Date"
STAFF_COLUMN = "Client Notes - Enrollment Level Staff Full Name"
NOTE_COLUMN = "Client Notes - Enrollment Level Note"
TIM_COLUMN = "TIM"
TEAM_COLUMN = "Team"

EXPECTED_COLUMNS = [
    UID_COLUMN,
    LAST_NAME_COLUMN,
    PROGRAM_COLUMN,
    START_COLUMN,
    EXIT_COLUMN,
    NOTE_DATE_COLUMN,
    ADDED_DATE_COLUMN,
    STAFF_COLUMN,
    NOTE_COLUMN,
]

DEDUPE_COLUMNS = [
    UID_COLUMN,
    EXIT_COLUMN,
    NOTE_DATE_COLUMN,
    ADDED_DATE_COLUMN,
    STAFF_COLUMN,
    NOTE_COLUMN,
]


class TimelinessService:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def process(self, program, uploaded_file, start_date, end_date, structure, current_metrics):
        program = str(program).strip().upper()
        if program not in {"MHRT", "RRT"}:
            raise ValueError("Program must be MHRT or RRT.")
        start = self._parse_date(start_date, "start")
        end = self._parse_date(end_date, "end")
        if start > end:
            raise ValueError("Start date must be on or before end date.")

        filename = Path(uploaded_file.filename).name
        self._validate_filename_program(program, filename)
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError("Case Notes upload must be a CSV, XLSX, or XLS file.")
        content = uploaded_file.read()
        if not content:
            raise ValueError("The uploaded Case Notes file is empty.")

        raw = self._load_frame(content, extension)
        clean = self._prepare_frame(raw, program, start, end)
        rows_before_dedupe = len(clean)
        clean = clean.drop_duplicates(subset=DEDUPE_COLUMNS, keep="first").copy()

        staff_to_team = {}
        for manager in structure.get("managers", []):
            for staff in manager.get("staff", []):
                staff_to_team[self._staff_key(staff)] = (
                    staff,
                    manager.get("name", "No Team") or "No Team",
                )

        def resolve_staff(value):
            cleaned = self._clean_text(value)
            direct = staff_to_team.get(self._staff_key(cleaned))
            if direct:
                return direct
            if "," in cleaned:
                last, first = [part.strip() for part in cleaned.split(",", 1)]
                reversed_match = staff_to_team.get(self._staff_key(f"{first} {last}"))
                if reversed_match:
                    return reversed_match
            return cleaned, "No Team"

        resolved = clean[STAFF_COLUMN].map(resolve_staff)
        clean[STAFF_COLUMN] = resolved.map(lambda item: item[0])
        clean[TEAM_COLUMN] = resolved.map(lambda item: item[1])

        report_rows = clean[clean[STAFF_COLUMN].ne("")].copy()
        counts = (
            report_rows.groupby([TEAM_COLUMN, STAFF_COLUMN], dropna=False)[UID_COLUMN]
            .count().rename("Count of Clients Unique Identifier").reset_index()
        )
        averages = (
            report_rows.dropna(subset=[TIM_COLUMN])
            .groupby([TEAM_COLUMN, STAFF_COLUMN], dropna=False)
            .agg(**{
                "Average of TIM": (TIM_COLUMN, "mean"),
                "_TIM Count": (TIM_COLUMN, "count"),
            }).reset_index()
        )

        metrics = {name: dict(values) for name, values in current_metrics.items()}
        roster_staff = [
            staff
            for manager in structure.get("managers", [])
            for staff in manager.get("staff", [])
        ]
        report_staff = list(dict.fromkeys(report_rows[STAFF_COLUMN].tolist()))
        for staff in dict.fromkeys(roster_staff + report_staff):
            metrics.setdefault(staff, {})
            metrics[staff]["case_notes"] = 0
            metrics[staff]["average_days"] = None
        for record in counts.to_dict("records"):
            metrics[record[STAFF_COLUMN]]["case_notes"] = int(record["Count of Clients Unique Identifier"])
        for record in averages.to_dict("records"):
            metrics[record[STAFF_COLUMN]]["average_days"] = round(float(record["Average of TIM"]), 2)

        roster_keys = {self._staff_key(staff) for staff in roster_staff}
        unmatched_staff = sorted(
            {staff for staff in report_staff if self._staff_key(staff) not in roster_keys},
            key=str.casefold,
        )
        output_name = self._write_workbook(program, raw, clean, counts, averages, start, end)
        return {
            "structure": structure,
            "metrics": metrics,
            "download_name": output_name,
            "source_filename": filename,
            "source_rows": int(len(raw)),
            "rows_in_date_range": int(rows_before_dedupe),
            "deduplicated_rows": int(len(clean)),
            "duplicates_removed": int(rows_before_dedupe - len(clean)),
            "negative_tim_rows": int(clean.attrs.get("negative_tim_rows", 0)),
            "staff_count": int(report_rows[STAFF_COLUMN].nunique()),
            "report_staff": report_staff,
            "unmatched_staff": unmatched_staff,
            "raw_frame": raw,
            "clean_frame": clean,
            "pivots": {"case_notes": counts, "average_days": averages},
        }

    def _load_frame(self, content, extension):
        stream = BytesIO(content)
        if extension == ".csv":
            try:
                frame = pd.read_csv(stream, dtype=str)
            except UnicodeDecodeError:
                stream.seek(0)
                frame = pd.read_csv(stream, dtype=str, encoding="latin-1")
            frame.columns = [self._clean_text(column) for column in frame.columns]
            return frame

        workbook = pd.ExcelFile(stream)
        for sheet_name in workbook.sheet_names:
            preview = pd.read_excel(workbook, sheet_name=sheet_name, header=None, nrows=30)
            for header_row in range(len(preview)):
                values = {self._clean_text(value) for value in preview.iloc[header_row].tolist() if pd.notna(value)}
                if {UID_COLUMN, NOTE_DATE_COLUMN, ADDED_DATE_COLUMN, STAFF_COLUMN}.issubset(values):
                    frame = pd.read_excel(workbook, sheet_name=sheet_name, header=header_row, dtype=str)
                    frame.columns = [self._clean_text(column) for column in frame.columns]
                    return frame
        raise ValueError("No Excel sheet contains the expected Case Notes columns.")

    def _prepare_frame(self, frame, program, start, end):
        frame = frame.copy()
        missing = [column for column in EXPECTED_COLUMNS if column not in frame.columns]
        if missing:
            raise ValueError("The Case Notes report is missing columns: " + ", ".join(missing))

        frame[NOTE_DATE_COLUMN] = pd.to_datetime(frame[NOTE_DATE_COLUMN], errors="coerce")
        frame[ADDED_DATE_COLUMN] = pd.to_datetime(frame[ADDED_DATE_COLUMN], errors="coerce")
        frame[START_COLUMN] = pd.to_datetime(frame[START_COLUMN], errors="coerce")
        frame[EXIT_COLUMN] = pd.to_datetime(frame[EXIT_COLUMN], errors="coerce")
        frame = frame[frame[NOTE_DATE_COLUMN].between(start, end, inclusive="both")].copy()

        has_program_values = frame[PROGRAM_COLUMN].fillna("").astype(str).str.strip().ne("").any()
        if has_program_values:
            dated_rows = len(frame)
            frame = frame[frame[PROGRAM_COLUMN].map(lambda value: self._matches_program(value, program))].copy()
            if dated_rows and frame.empty:
                raise ValueError(f"The Case Notes report has no {program} rows in the selected date range.")

        for column in [UID_COLUMN, LAST_NAME_COLUMN, PROGRAM_COLUMN, STAFF_COLUMN, NOTE_COLUMN]:
            frame[column] = frame[column].fillna("").astype(str).map(self._clean_text)
        frame[TIM_COLUMN] = (frame[ADDED_DATE_COLUMN] - frame[NOTE_DATE_COLUMN]).dt.total_seconds() / 86400
        negative_tim_rows = int(frame[TIM_COLUMN].lt(0).sum())
        frame.loc[frame[TIM_COLUMN] < 0, TIM_COLUMN] = pd.NA
        frame.attrs["negative_tim_rows"] = negative_tim_rows
        return frame.reset_index(drop=True)

    def _write_workbook(self, program, raw, clean, counts, averages, start, end):
        program_dir = self.output_dir / program
        program_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{program}_Case_Notes_{start:%Y%m%d}_{end:%Y%m%d}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        workbook = Workbook()
        pivot_sheet = workbook.active
        pivot_sheet.title = "Pivots"
        self.write_pivots(pivot_sheet, counts, averages, start_row=1, start_column=1)
        self._write_frame(workbook, "Case_Notes_Raw", raw)
        self._write_frame(workbook, "Case_Notes_Clean", clean)
        workbook.save(program_dir / filename)
        return filename

    @staticmethod
    def write_pivots(sheet, counts, averages, start_row=1, start_column=1):
        sections = [
            ("Case Notes (Enrollment Level)", counts, "Count of Clients Unique Identifier"),
            ("Average # of Days between Case Note Date and Date Added", averages, "Average of TIM"),
        ]
        row = start_row
        for title, frame, value_column in sections:
            sheet.merge_cells(
                start_row=row,
                start_column=start_column,
                end_row=row,
                end_column=start_column + 2,
            )
            sheet.cell(row, start_column, title).font = Font(bold=True)
            sheet.cell(row, start_column).fill = PatternFill("solid", fgColor="B4C6E7")
            sheet.cell(row, start_column).alignment = Alignment(wrap_text=True, vertical="center")
            sheet.row_dimensions[row].height = 30
            row += 1
            headers = [TEAM_COLUMN, STAFF_COLUMN, value_column]
            for offset, header in enumerate(headers):
                cell = sheet.cell(row, start_column + offset, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            sheet.row_dimensions[row].height = 45
            row += 1
            for record in frame.to_dict("records"):
                sheet.cell(row, start_column, record[TEAM_COLUMN])
                sheet.cell(row, start_column + 1, record[STAFF_COLUMN])
                value = record[value_column]
                sheet.cell(row, start_column + 2, round(float(value), 2) if value_column == "Average of TIM" else int(value))
                if value_column == "Average of TIM":
                    sheet.cell(row, start_column + 2).number_format = "0.00"
                row += 1
            total = sheet.cell(row, start_column, "Grand Total")
            total.font = Font(bold=True)
            total.fill = PatternFill("solid", fgColor="E2F0D9")
            if value_column == "Count of Clients Unique Identifier":
                value = int(frame[value_column].sum()) if not frame.empty else 0
            else:
                if not frame.empty and "_TIM Count" in frame.columns and frame["_TIM Count"].sum():
                    value = round(
                        float((frame[value_column] * frame["_TIM Count"]).sum() / frame["_TIM Count"].sum()),
                        2,
                    )
                else:
                    value = 0
            sheet.cell(row, start_column + 2, value).font = Font(bold=True)
            sheet.cell(row, start_column + 2).fill = PatternFill("solid", fgColor="E2F0D9")
            if value_column == "Average of TIM":
                sheet.cell(row, start_column + 2).number_format = "0.00"
            row += 2
        for column, width in enumerate([28, 52, 30], start=start_column):
            sheet.column_dimensions[get_column_letter(column)].width = width
        return row

    @staticmethod
    def _write_frame(workbook, title, frame):
        sheet = workbook.create_sheet(title)
        for column_index, column in enumerate(frame.columns, start=1):
            cell = sheet.cell(1, column_index, str(column))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(values, start=1):
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                sheet.cell(row_index, column_index, value)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for index, column in enumerate(frame.columns, start=1):
            width = min(max(14, len(str(column)) + 2), 45)
            sheet.column_dimensions[get_column_letter(index)].width = width

    @staticmethod
    def _parse_date(value, label):
        try:
            return pd.Timestamp(value).normalize()
        except (TypeError, ValueError):
            raise ValueError(f"Select a valid {label} date.") from None

    @staticmethod
    def _clean_text(value):
        if pd.isna(value):
            return ""
        return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()

    @classmethod
    def _normalized(cls, value):
        return cls._clean_text(value).casefold()

    @classmethod
    def _staff_key(cls, value):
        cleaned = re.sub(r"\(\*?\s*denotes inactive\)", "", cls._clean_text(value), flags=re.IGNORECASE)
        if cleaned.casefold() == "jayne lee":
            cleaned = "Jayna Lee"
        return re.sub(r"[^a-z0-9]+", "", cleaned.casefold())

    @staticmethod
    def _matches_program(value, program):
        text = re.sub(r"[^a-z0-9]+", " ", str(value).casefold()).strip()
        tokens = set(text.split())
        if program == "MHRT":
            return "mhrt" in tokens or "mobile homelessness resolution" in text
        return "rrt" in tokens or "rapid response team" in text

    @staticmethod
    def _validate_filename_program(program, filename):
        normalized = re.sub(r"[^a-z0-9]+", " ", filename.casefold()).strip()
        tokens = set(normalized.split())
        is_mhrt = "mhrt" in tokens or "mobile homelessness resolution" in normalized
        is_rrt = "rrt" in tokens or "rapid response team" in normalized
        if program == "MHRT" and is_rrt and not is_mhrt:
            raise ValueError("This appears to be an RRT Case Notes report. Switch the app to RRT first.")
        if program == "RRT" and is_mhrt and not is_rrt:
            raise ValueError("This appears to be an MHRT Case Notes report. Switch the app to MHRT first.")
