from datetime import datetime
from io import BytesIO
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


UID_COLUMNS = ["Unique ID", "Clients Unique Identifier"]
DATE_COLUMNS = ["Date of Contact", "CLS Date", "Current Living Situation Date Created Date"]
STAFF_COLUMNS = ["Staff Created", "Assigned Staff (* denotes Inactive)"]
TEAM_COLUMNS = ["team", "Team"]
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


class CLSService:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def process(self, program, uploaded_file, start_date, end_date, structure, current_metrics, team_uploaded_file=None):
        program = str(program).strip().upper()
        if program not in {"MHRT", "RRT"}:
            raise ValueError("Program must be MHRT or RRT.")
        start = self._parse_date(start_date, "start")
        end = self._parse_date(end_date, "end")
        if start > end:
            raise ValueError("Start date must be on or before end date.")

        filename = Path(uploaded_file.filename).name
        self._validate_filename_program(program, filename)
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError("CLS upload must be a CSV, XLSX, or XLS file.")
        content = uploaded_file.read()
        if not content:
            raise ValueError("The uploaded CLS file is empty.")

        instance_source, unique_source = self._load_cls_frames(content, extension)
        frame, uid_col, date_col, staff_col = self._prepare_cls_frame(instance_source, start, end)
        unique_frame, unique_uid_col, unique_date_col, unique_staff_col = self._prepare_cls_frame(
            unique_source, start, end
        )
        unique_frame = unique_frame.sort_values(unique_date_col, ascending=False)
        # A client is unique for the whole report, not once per staff member.
        # Keeping the newest row mirrors the pandas-created Unique worksheet.
        unique_frame = unique_frame.drop_duplicates(subset=[unique_uid_col], keep="first").copy()

        total_summary = (
            frame.groupby(staff_col, dropna=False)
            .agg(cls_assessments=(uid_col, "size"))
            .reset_index()
            .rename(columns={staff_col: "staff"})
        )
        unique_summary = (
            unique_frame.groupby(unique_staff_col, dropna=False)
            .agg(clients_served=(unique_uid_col, "size"))
            .reset_index()
            .rename(columns={unique_staff_col: "staff"})
        )
        summary = total_summary.merge(unique_summary, on="staff", how="outer").fillna(0)
        summary[["cls_assessments", "clients_served"]] = summary[["cls_assessments", "clients_served"]].astype(int)
        team_col = next((column for column in TEAM_COLUMNS if column in frame.columns), None)
        team_lookup = {}
        if team_col:
            team_lookup = (
                frame[[staff_col, team_col]].dropna(subset=[team_col]).drop_duplicates(subset=[staff_col])
                .set_index(staff_col)[team_col].astype(str).to_dict()
            )
        team_file_used = False
        if team_uploaded_file and team_uploaded_file.filename:
            team_result = self.structure_from_team_upload(program, team_uploaded_file)
            structure = team_result["structure"]
            team_lookup = self._metric_team_lookup(
                frame, staff_col, team_result["staff_to_manager"], team_lookup
            )
            team_file_used = True

        metrics = {name: dict(values) for name, values in current_metrics.items()}
        for row in summary.to_dict("records"):
            metrics.setdefault(row["staff"], {})
            metrics[row["staff"]]["cls_assessments"] = int(row["cls_assessments"])
            metrics[row["staff"]]["clients_served"] = int(row["clients_served"])

        roster_staff = {
            staff.casefold()
            for manager in structure.get("managers", [])
            for staff in manager.get("staff", [])
        }
        unmatched_staff = sorted(
            {staff for staff in summary["staff"].tolist() if staff.casefold() not in roster_staff},
            key=str.casefold,
        )
        filtered_output = self._with_team_column(frame, staff_col, team_lookup)
        unique_output = self._with_team_column(unique_frame, unique_staff_col, team_lookup)
        output_name = self._write_summary(
            program,
            filtered_output,
            unique_output,
            uid_col,
            unique_uid_col,
            staff_col,
            unique_staff_col,
            start,
            end,
        )
        return {
            "structure": structure,
            "metrics": metrics,
            "download_name": output_name,
            "rows_in_date_range": int(len(frame)),
            "unique_clients": int(unique_frame[unique_uid_col].nunique()),
            "staff_count": int(summary["staff"].nunique()),
            "report_staff": summary["staff"].tolist(),
            "unmatched_staff": unmatched_staff,
            "source_filename": filename,
            "team_file_used": team_file_used,
        }

    def _load_cls_frames(self, content, extension):
        stream = BytesIO(content)
        if extension == ".csv":
            try:
                frame = pd.read_csv(stream)
            except UnicodeDecodeError:
                stream.seek(0)
                frame = pd.read_csv(stream, encoding="latin-1")
            return frame, frame.copy()

        workbook = pd.ExcelFile(stream)
        candidates = []
        for sheet in workbook.sheet_names:
            candidate = pd.read_excel(workbook, sheet_name=sheet)
            candidate.columns = [str(column).strip() for column in candidate.columns]
            has_uid = any(column in candidate.columns for column in UID_COLUMNS)
            has_date = any(column in candidate.columns for column in DATE_COLUMNS)
            has_staff = any(column in candidate.columns for column in STAFF_COLUMNS)
            if has_uid and has_date and has_staff:
                candidates.append((sheet, candidate))
        if not candidates:
            raise ValueError("No Excel sheet contains the expected client ID, CLS date, and staff columns.")
        if len(candidates) == 1:
            return candidates[0][1], candidates[0][1].copy()

        unique_candidates = [item for item in candidates if "unique" in item[0].casefold()]
        other_candidates = [item for item in candidates if "unique" not in item[0].casefold()]
        unique_frame = unique_candidates[0][1] if unique_candidates else candidates[0][1]
        instance_frame = other_candidates[0][1] if other_candidates else candidates[0][1]
        return instance_frame, unique_frame

    def _prepare_cls_frame(self, frame, start, end):
        frame = frame.copy()
        frame.columns = [str(column).strip() for column in frame.columns]
        uid_col = self._first_column(frame, UID_COLUMNS, "client ID")
        date_col = self._first_column(frame, DATE_COLUMNS, "CLS date")
        staff_col = self._first_column(frame, STAFF_COLUMNS, "staff")
        frame[date_col] = pd.to_datetime(frame[date_col], errors="coerce")
        frame = frame[frame[date_col].between(start, end, inclusive="both")].copy()
        frame[staff_col] = frame[staff_col].map(self._clean_staff_key)
        frame[uid_col] = frame[uid_col].fillna("").astype(str).str.strip()
        frame = frame[(frame[staff_col] != "") & (frame[uid_col] != "")]
        return frame, uid_col, date_col, staff_col

    def structure_from_team_upload(self, program, uploaded_file):
        program = str(program).strip().upper()
        if program not in {"MHRT", "RRT"}:
            raise ValueError("Program must be MHRT or RRT.")
        filename = Path(uploaded_file.filename).name
        self._validate_team_filename_program(program, filename)
        extension = Path(filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError("Team mapping must be a CSV, XLSX, or XLS file.")
        content = uploaded_file.read()
        if not content:
            raise ValueError("The uploaded Team file is empty.")

        team_frame = self._load_first_table(content, extension)
        team_frame.columns = [str(column).strip() for column in team_frame.columns]
        if len(team_frame.columns) < 3:
            raise ValueError(
                "Team file must contain staff/full name, team/manager, and optionally Services User Creating."
            )

        columns = list(team_frame.columns)
        program_col = next((column for column in columns if "program" in column.casefold()), None)
        manager_col = next(
            (column for column in columns if "team" in column.casefold() or "manager" in column.casefold()),
            columns[-1],
        )
        service_user_col = next(
            (
                column for column in columns
                if "username" in column.casefold()
                or ("service" in column.casefold() and "user" in column.casefold())
                or ("service" in column.casefold() and "creating" in column.casefold())
            ),
            None,
        )
        fullname_col = next(
            (
                column for column in columns
                if "fullname" in column.casefold()
                or "full name" in column.casefold()
                or column.casefold().strip() == "member"
            ),
            None,
        )
        active_col = next(
            (column for column in columns if column.casefold().strip() == "active"),
            None,
        )

        alias_layout = service_user_col is not None and fullname_col is not None
        if not alias_layout and program_col is None:
            # Legacy Team files use column 1 for staff and column 3 for team.
            # Column 2 may contain a short manager label such as Cory or Autumn.
            fullname_col = columns[0]
            manager_col = columns[2]
        elif not alias_layout:
            fullname_col = next(
                (column for column in columns if column not in {program_col, manager_col}),
                columns[0],
            )

        if program_col:
            selected = team_frame[program_col].fillna("").astype(str).str.strip().str.upper()
            team_frame = team_frame[selected == program].copy()
            if team_frame.empty:
                raise ValueError(f"The Team file has no {program} rows.")

        team_frame.loc[:, fullname_col] = team_frame[fullname_col].map(self._clean_staff_key)
        team_frame.loc[:, manager_col] = team_frame[manager_col].fillna("").astype(str).str.strip()
        if alias_layout:
            team_frame.loc[:, service_user_col] = team_frame[service_user_col].map(self._clean_staff_key)
        if active_col:
            team_frame.loc[:, active_col] = team_frame[active_col].map(self._active_value)
        team_frame = team_frame[
            (team_frame[fullname_col] != "")
            & (team_frame[manager_col] != "")
            & (team_frame[manager_col] != "0")
        ]
        if alias_layout:
            duplicate_username = team_frame[service_user_col].duplicated(keep="first")
            team_frame = team_frame[(team_frame[service_user_col] == "") | ~duplicate_username]
        team_frame = team_frame.drop_duplicates(subset=[fullname_col], keep="first")
        if team_frame.empty:
            raise ValueError("The Team file contains no usable staff-to-manager rows.")

        managers = []
        managers_by_key = {}
        staff_to_manager = {}
        service_user_to_staff = {}
        staff_to_service_user = {}
        active = {}
        selected_columns = (
            [fullname_col, manager_col]
            + ([service_user_col] if alias_layout else [])
            + ([active_col] if active_col else [])
        )
        for row in team_frame[selected_columns].to_dict("records"):
            staff = row[fullname_col]
            manager_name = row[manager_col]
            manager_key = manager_name.casefold()
            if manager_key not in managers_by_key:
                manager = {"name": manager_name, "staff": []}
                managers.append(manager)
                managers_by_key[manager_key] = manager
            managers_by_key[manager_key]["staff"].append(staff)
            staff_to_manager[staff.casefold()] = manager_name
            active[staff] = bool(row[active_col]) if active_col else True
            if alias_layout:
                service_user = row[service_user_col]
                if service_user:
                    service_user_to_staff[service_user.casefold()] = staff
                    staff_to_service_user[staff] = service_user

        return {
            "structure": {"program": program, "managers": managers, "active": active},
            "staff_to_manager": staff_to_manager,
            "service_user_to_staff": service_user_to_staff,
            "staff_to_service_user": staff_to_service_user,
            "staff_count": int(len(team_frame)),
            "manager_count": int(len(managers)),
            "source_filename": filename,
        }

    @staticmethod
    def _validate_team_filename_program(program, filename):
        normalized = re.sub(r"[^a-z0-9]+", " ", filename.casefold()).strip()
        tokens = set(normalized.split())
        mh_rt_file = "mhrt" in tokens or "mobile homelessness resolution" in normalized
        rrt_file = "rrt" in tokens or "rapid response team" in normalized
        if program == "MHRT" and rrt_file and not mh_rt_file:
            raise ValueError("This appears to be an RRT Team mapping. Switch the app to RRT before importing it.")
        if program == "RRT" and mh_rt_file and not rrt_file:
            raise ValueError("This appears to be an MHRT Team mapping. Switch the app to MHRT before importing it.")

    @staticmethod
    def _active_value(value):
        if isinstance(value, bool):
            return value
        return str(value).strip().casefold() not in {"no", "n", "false", "0", "inactive"}

    @staticmethod
    def _first_column(frame, candidates, label):
        for column in candidates:
            if column in frame.columns:
                return column
        raise ValueError(f"The CLS report is missing a recognized {label} column.")

    @staticmethod
    def _parse_date(value, label):
        try:
            return pd.Timestamp(value).normalize()
        except (TypeError, ValueError):
            raise ValueError(f"Select a valid {label} date.") from None

    @staticmethod
    def _validate_filename_program(program, filename):
        """Reject a report whose filename explicitly identifies the other program.

        Generic filenames remain valid because some users rename exports before upload.
        """
        normalized = re.sub(r"[^a-z0-9]+", " ", filename.casefold()).strip()
        mh_rt_file = "mhrt" in normalized.split() or "mobile homelessness resolution" in normalized
        rrt_file = "rrt" in normalized.split() or "rapid response team" in normalized
        if program == "MHRT" and rrt_file and not mh_rt_file:
            raise ValueError("This appears to be an RRT CLS report. Switch the app to RRT before processing it.")
        if program == "RRT" and mh_rt_file and not rrt_file:
            raise ValueError("This appears to be an MHRT CLS report. Switch the app to MHRT before processing it.")

    def _metric_team_lookup(self, report, metric_staff_col, mapping, fallback):
        report = report.copy()
        report["_metric_staff_key"] = report[metric_staff_col].map(self._clean_staff_key).str.casefold()
        report["_uploaded_team"] = report["_metric_staff_key"].map(mapping).fillna("")
        if "Assigned Staff (* denotes Inactive)" in report.columns:
            report["_assigned_staff_key"] = (
                report["Assigned Staff (* denotes Inactive)"].map(self._clean_staff_key).str.casefold()
            )
            report["_uploaded_team"] = report["_uploaded_team"].mask(
                report["_uploaded_team"] == "",
                report["_assigned_staff_key"].map(mapping).fillna(""),
            )
        uploaded_lookup = (
            report[report["_uploaded_team"] != ""]
            .drop_duplicates(subset=[metric_staff_col], keep="first")
            .set_index(metric_staff_col)["_uploaded_team"]
            .astype(str)
            .to_dict()
        )
        return {**fallback, **uploaded_lookup}

    def _load_first_table(self, content, extension):
        stream = BytesIO(content)
        if extension == ".csv":
            try:
                return pd.read_csv(stream)
            except UnicodeDecodeError:
                stream.seek(0)
                return pd.read_csv(stream, encoding="latin-1")
        workbook = pd.ExcelFile(stream)
        if not workbook.sheet_names:
            raise ValueError("The Team workbook has no worksheets.")
        return pd.read_excel(workbook, sheet_name=workbook.sheet_names[0])

    @staticmethod
    def _with_team_column(frame, staff_col, team_lookup):
        output = frame.copy()
        existing_team_col = next((column for column in TEAM_COLUMNS if column in output.columns), None)
        existing = (
            output[existing_team_col].fillna("").astype(str).str.strip()
            if existing_team_col
            else pd.Series("", index=output.index, dtype="object")
        )
        mapped = output[staff_col].map(team_lookup).fillna("").astype(str).str.strip()
        output["Team"] = mapped.mask(mapped == "", existing)
        if existing_team_col and existing_team_col != "Team":
            output = output.drop(columns=[existing_team_col])
        return output

    @staticmethod
    def _clean_staff_key(value):
        text = str(value if pd.notna(value) else "").replace("\u00a0", " ")
        text = re.sub(r"Deactivated per ticket.*", "", text)
        cleaned = re.sub(r"\s+", " ", text).strip()
        return "Jayna Lee" if cleaned.casefold() == "jayne lee" else cleaned

    @staticmethod
    def _add_missing_staff(structure, report_staff, team_lookup, reorganize=False):
        structure = {
            "program": structure.get("program", ""),
            "active": dict(structure.get("active", {})),
            "managers": [
                {"name": manager.get("name", ""), "staff": list(manager.get("staff", []))}
                for manager in structure.get("managers", [])
            ],
        }
        report_keys = {staff.casefold() for staff in report_staff}
        if reorganize:
            for manager in structure["managers"]:
                manager["staff"] = [staff for staff in manager["staff"] if staff.casefold() not in report_keys]
            structure["managers"] = [manager for manager in structure["managers"] if manager["staff"]]
        known = {staff.casefold() for manager in structure["managers"] for staff in manager["staff"]}
        managers_by_name = {manager["name"].casefold(): manager for manager in structure["managers"]}
        for staff in report_staff:
            if staff.casefold() in known:
                continue
            manager_name = str(team_lookup.get(staff, "Imported Staff")).strip() or "Imported Staff"
            manager = managers_by_name.get(manager_name.casefold())
            if manager is None:
                manager = {"name": manager_name, "staff": []}
                structure["managers"].append(manager)
                managers_by_name[manager_name.casefold()] = manager
            manager["staff"].append(staff)
            structure["active"].setdefault(staff, True)
            known.add(staff.casefold())
        return structure

    def _write_summary(
        self,
        program,
        filtered_frame,
        unique_frame,
        filtered_uid_col,
        unique_uid_col,
        filtered_staff_col,
        unique_staff_col,
        start,
        end,
    ):
        program_dir = self.output_dir / program
        program_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{program}_CLS_Pivots_{start:%Y%m%d}_{end:%Y%m%d}_{timestamp}.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_grouped_summary(
            workbook, "Unique Summary", f"{program} Unique", unique_frame,
            unique_uid_col, unique_staff_col,
        )
        self._write_grouped_summary(
            workbook, "Summary", f"{program} Total", filtered_frame,
            filtered_uid_col, filtered_staff_col,
        )
        self._write_data_sheet(workbook, "Filtered", filtered_frame)
        self._write_data_sheet(workbook, "Unique", unique_frame)
        workbook.save(program_dir / filename)
        return filename

    @staticmethod
    def _write_grouped_summary(workbook, sheet_name, title, frame, uid_col, staff_col):
        sheet = workbook.create_sheet(sheet_name)
        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True)
        sheet["A3"] = "Row Labels"
        sheet["B3"] = f"Count of {uid_col}"
        for cell in sheet[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B4C6E7")
            cell.border = Border(bottom=Side(style="thin", color="7F8FA6"))

        grouped = (
            frame.assign(
                _team=frame["Team"].fillna("").astype(str).str.strip().replace("", "(blank)"),
                _staff=frame[staff_col].fillna("").astype(str).str.strip(),
            )
            .groupby(["_team", "_staff"], dropna=False)
            .size()
            .reset_index(name="count")
        )
        grouped["_team_sort"] = grouped["_team"].map(
            lambda value: (value == "(blank)", value.casefold())
        )
        grouped["_staff_sort"] = grouped["_staff"].str.casefold()
        grouped = grouped.sort_values(["_team_sort", "_staff_sort"], kind="stable")

        row_number = 4
        for team, team_rows in grouped.groupby("_team", sort=False):
            team_total = int(team_rows["count"].sum())
            sheet.cell(row_number, 1, team)
            sheet.cell(row_number, 2, team_total)
            for cell in sheet[row_number]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="BFBFBF")
                cell.border = Border(bottom=Side(style="thin", color="7F8FA6"))
            row_number += 1
            for record in team_rows.to_dict("records"):
                sheet.cell(row_number, 1, record["_staff"])
                sheet.cell(row_number, 1).alignment = Alignment(indent=2)
                sheet.cell(row_number, 2, int(record["count"]))
                row_number += 1

        total_row = row_number
        sheet.cell(total_row, 1, "Grand Total")
        sheet.cell(total_row, 2, int(len(frame)))
        for cell in sheet[total_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="B4C6E7")
            cell.border = Border(top=Side(style="thin", color="7F8FA6"))
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 20
        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A3:B{total_row}"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def _write_data_sheet(workbook, sheet_name, frame):
        sheet = workbook.create_sheet(sheet_name)
        columns = list(frame.columns)
        for column_number, header in enumerate(columns, start=1):
            cell = sheet.cell(1, column_number, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for row_number, values in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_number, value in enumerate(values, start=1):
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                cell = sheet.cell(row_number, column_number, value)
                if isinstance(value, datetime):
                    cell.number_format = "yyyy-mm-dd"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(frame) + 1)}"
        if len(frame):
            table = Table(
                displayName=f"{sheet_name}Data",
                ref=f"A1:{get_column_letter(len(columns))}{len(frame) + 1}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            sheet.add_table(table)
        for column_number, header in enumerate(columns, start=1):
            values = [str(header)] + [str(value) for value in frame.iloc[:, column_number - 1].dropna().head(200)]
            sheet.column_dimensions[get_column_letter(column_number)].width = min(max(map(len, values)) + 2, 42)
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
