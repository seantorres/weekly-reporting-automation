import json
import re
import csv
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROGRAMS = {"MHRT", "RRT"}
NAME_CORRECTIONS = {"jayne lee": "Jayna Lee"}
METRIC_FIELDS = {
    "clients_served",
    "services",
    "referrals_to_shelter",
    "currently_enrolled",
    "engaged",
    "exits",
    "positive_exits",
    "ph_exits",
    "no_exit_interview",
    "chronically_homeless",
    "first_enrollment",
    "cls_assessments",
    "case_notes",
    "average_days",
}
NULLABLE_METRIC_FIELDS = {"average_days"}


class StructureService:
    def __init__(self, data_dir: Path, default_mapping_dir: Path | None = None):
        self.data_dir = data_dir
        self.default_mapping_dir = default_mapping_dir
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def _program(self, program: str) -> str:
        normalized = program.strip().upper()
        if normalized not in PROGRAMS:
            raise ValueError("Program must be MHRT or RRT.")
        return normalized

    def _read_json(self, path: Path, default):
        if not path.exists():
            return default
        with path.open("r", encoding="utf-8") as file:
            return json.load(file)

    def _write_json(self, path: Path, payload):
        temporary = path.with_suffix(".tmp")
        with temporary.open("w", encoding="utf-8") as file:
            json.dump(payload, file, indent=2)
        temporary.replace(path)

    def load_structure(self, program: str):
        program = self._program(program)
        default = {"program": program, "managers": [], "active": {}}
        saved = self._read_json(self.data_dir / f"{program.lower()}_structure.json", default)
        return self._clean_structure(program, saved)

    def save_structure(self, program: str, payload):
        program = self._program(program)
        result = self._clean_structure(program, payload)
        self._write_json(self.data_dir / f"{program.lower()}_structure.json", result)
        if isinstance(payload.get("service_users"), dict):
            self.save_service_users(program, payload["service_users"])
        return result

    def _clean_structure(self, program, payload):
        managers = payload.get("managers")
        if not isinstance(managers, list):
            raise ValueError("The managers field must be a list.")

        cleaned = []
        seen_staff = set()
        supplied_active = payload.get("active", {})
        if not isinstance(supplied_active, dict):
            supplied_active = {}
        active_by_key = {
            self._clean_name(staff).casefold(): self._active_value(value)
            for staff, value in supplied_active.items()
            if self._clean_name(staff)
        }
        active = {}
        for manager in managers:
            name = self._clean_name(manager.get("name", ""))
            if not name:
                continue
            staff_list = []
            for staff in manager.get("staff", []):
                staff_name = self._clean_name(staff)
                key = staff_name.casefold()
                if staff_name and key not in seen_staff:
                    staff_list.append(staff_name)
                    seen_staff.add(key)
                    active[staff_name] = active_by_key.get(key, True)
            cleaned.append({"name": name, "staff": staff_list})

        result = {"program": program, "managers": cleaned, "active": active}
        return result

    def export_team_mapping(self, program, structure, service_users, output_dir):
        program = self._program(program)
        structure = self._clean_structure(program, structure)
        output_dir = Path(output_dir) / program
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{program}_Team_Mapping_{timestamp}.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Team Mapping"
        headers = ["Program", "Member", "Manager", "Services User Creating", "Active"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(horizontal="center")

        active = structure.get("active", {})
        for manager in structure.get("managers", []):
            for staff in manager.get("staff", []):
                sheet.append([
                    program,
                    staff,
                    manager.get("name", ""),
                    service_users.get(staff, ""),
                    "Yes" if active.get(staff, True) else "No",
                ])
        sheet.freeze_panes = "A2"
        widths = {"A": 12, "B": 28, "C": 24, "D": 28, "E": 12}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(output_dir / filename)
        return filename

    def list_presets(self, program: str):
        program = self._program(program)
        presets = self._read_json(self.data_dir / "staff_presets.json", {})
        return sorted(presets.get(program, {}).keys(), key=str.casefold)

    def save_preset(self, program: str, name: str, structure):
        program = self._program(program)
        name = self._clean_name(name)
        if not name:
            raise ValueError("Enter a preset name.")
        cleaned = self._clean_structure(program, structure)
        service_users = self.save_service_users(
            program,
            structure.get("service_users", self.load_service_users(program)),
        )
        presets = self._read_json(self.data_dir / "staff_presets.json", {})
        presets.setdefault(program, {})[name] = {
            "structure": cleaned,
            "service_users": service_users,
        }
        self._write_json(self.data_dir / "staff_presets.json", presets)
        self._write_json(self.data_dir / f"{program.lower()}_structure.json", cleaned)
        return {
            "name": name,
            "structure": cleaned,
            "service_users": service_users,
            "presets": self.list_presets(program),
        }

    def load_preset(self, program: str, name: str):
        program = self._program(program)
        name = self._clean_name(name)
        presets = self._read_json(self.data_dir / "staff_presets.json", {})
        saved = presets.get(program, {}).get(name)
        if saved is None:
            raise ValueError("Choose a saved preset.")
        if isinstance(saved, dict) and "structure" in saved:
            structure = saved["structure"]
            service_users = saved.get("service_users", {})
        else:
            structure = saved
            service_users = self.load_service_users(program)
        cleaned = self._clean_structure(program, structure)
        self._write_json(self.data_dir / f"{program.lower()}_structure.json", cleaned)
        service_users = self.save_service_users(program, service_users)
        return {"name": name, "structure": cleaned, "service_users": service_users}

    def load_service_users(self, program: str):
        program = self._program(program)
        defaults = self._load_default_service_users(program)
        saved = self._read_json(self.data_dir / f"{program.lower()}_service_users.json", {})
        if not isinstance(saved, dict):
            saved = {}

        # Older builds displayed an unknown username as both the staff name and
        # username. Do not let those self-links override a bundled full-name
        # match such as MCano -> Maria Cano.
        default_by_username = {
            str(username).strip().casefold(): staff
            for staff, username in defaults.items()
        }
        merged = dict(defaults)
        for staff, username in saved.items():
            staff_name = self._clean_name(staff)
            user_name = self._clean_name(username)
            known_staff = default_by_username.get(user_name.casefold())
            if known_staff and staff_name.casefold() == user_name.casefold():
                continue
            if known_staff and known_staff.casefold() != staff_name.casefold():
                merged.pop(known_staff, None)
            merged[staff_name] = user_name
        return merged

    def _load_default_service_users(self, program):
        if self.default_mapping_dir is None:
            return {}
        mapping_path = self.default_mapping_dir / f"{program}_service_user_mapping.csv"
        if not mapping_path.exists():
            return {}
        mappings = {}
        with mapping_path.open("r", encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                staff = self._clean_name(row.get("Fullname", ""))
                username = self._clean_name(row.get("Services User Creating", ""))
                if staff and username:
                    mappings[staff] = username
        return mappings

    def load_service_user_teams(self, program):
        program = self._program(program)
        if self.default_mapping_dir is None:
            return {}
        mapping_path = self.default_mapping_dir / f"{program}_service_user_mapping.csv"
        if not mapping_path.exists():
            return {}
        teams = {}
        with mapping_path.open("r", encoding="utf-8-sig", newline="") as file:
            for row in csv.DictReader(file):
                staff = self._clean_name(row.get("Fullname", ""))
                team = self._clean_name(row.get("Team", ""))
                if staff and team:
                    teams[staff] = team
        return teams

    def save_service_users(self, program: str, payload):
        program = self._program(program)
        if not isinstance(payload, dict):
            raise ValueError("Service-user mappings must be keyed by staff name.")
        cleaned = {}
        usernames = set()
        for staff, username in payload.items():
            staff_name = self._clean_name(staff)
            user_name = self._clean_name(username)
            if not staff_name or not user_name:
                continue
            user_key = user_name.casefold()
            if user_key in usernames:
                raise ValueError(f"Services User Creating '{user_name}' is assigned more than once.")
            cleaned[staff_name] = user_name
            usernames.add(user_key)
        self._write_json(self.data_dir / f"{program.lower()}_service_users.json", cleaned)
        return cleaned

    def merge_service_users(self, program: str, payload):
        merged = self.load_service_users(program)
        merged.update(payload or {})
        return self.save_service_users(program, merged)

    def load_metrics(self, program: str):
        program = self._program(program)
        saved = self._read_json(self.data_dir / f"{program.lower()}_metrics.json", {})
        cleaned = {}
        for staff, values in saved.items():
            staff_name = self._clean_name(staff)
            if staff_name and isinstance(values, dict):
                cleaned.setdefault(staff_name, {}).update(values)
        return cleaned

    def save_metrics(self, program: str, payload):
        program = self._program(program)
        if not isinstance(payload, dict):
            raise ValueError("Metrics must be an object keyed by staff name.")

        cleaned = {}
        for staff, metrics in payload.items():
            staff_name = self._clean_name(staff)
            if not staff_name or not isinstance(metrics, dict):
                continue
            cleaned[staff_name] = {
                field: (
                    None
                    if field in NULLABLE_METRIC_FIELDS
                    and metrics.get(field) in {None, "", "-"}
                    else self._nonnegative_number(metrics.get(field, 0))
                )
                for field in METRIC_FIELDS
            }

        self._write_json(self.data_dir / f"{program.lower()}_metrics.json", cleaned)
        return cleaned

    def install_mock_data(self, program: str):
        program = self._program(program)
        if program != "MHRT":
            raise ValueError("The included preview dataset is for MHRT.")

        structure = {
            "program": "MHRT",
            "managers": [
                {
                    "name": "Sample Manager A",
                    "staff": ["Alex Rivera", "Jordan Lee", "Taylor Morgan", "Casey Nguyen"],
                },
                {
                    "name": "Sample Manager B",
                    "staff": ["Morgan Ellis", "Riley Carter", "Cameron Diaz", "Avery Brooks"],
                },
                {
                    "name": "Sample Manager C",
                    "staff": ["Parker Reed", "Jamie Patel", "Quinn Foster", "Drew Kim"],
                },
            ],
        }

        rows = [
            (15, 65, 21, 37, 14, 6, 0, 0, 6, 42, 4, 29, 25, 0),
            (5, 35, 8, 9, 3, 2, 1, 1, 1, 9, 3, 9, 16, 2.25),
            (8, 28, 3, 38, 14, 0, 0, 0, 0, 30, 5, 13, 13, 0.08),
            (12, 44, 5, 31, 18, 3, 1, 0, 2, 25, 4, 21, 18, 1.2),
            (14, 32, 15, 32, 16, 0, 0, 0, 0, 15, 3, 34, 17, 0),
            (17, 59, 1, 39, 34, 4, 2, 2, 0, 36, 14, 30, 31, 0),
            (3, 6, 2, 40, 18, 0, 0, 0, 0, 33, 10, 4, 2, 0.5),
            (5, 6, 0, 16, 11, 0, 0, 0, 0, 14, 3, 6, 10, 6.33),
            (10, 30, 0, 25, 5, 0, 0, 0, 0, 23, 2, 21, 20, 3.39),
            (4, 10, 0, 23, 5, 0, 0, 0, 0, 12, 7, 4, 5, 6.25),
            (14, 32, 3, 32, 15, 0, 0, 0, 0, 23, 5, 24, 20, 0.7),
            (12, 27, 2, 42, 32, 1, 0, 0, 1, 39, 4, 14, 17, 0),
        ]
        fields = [
            "clients_served", "services", "referrals_to_shelter",
            "currently_enrolled", "engaged", "exits", "positive_exits",
            "ph_exits", "no_exit_interview", "chronically_homeless",
            "first_enrollment", "cls_assessments", "case_notes", "average_days",
        ]
        staff_names = [staff for manager in structure["managers"] for staff in manager["staff"]]
        metrics = {
            staff: dict(zip(fields, values))
            for staff, values in zip(staff_names, rows)
        }

        self._write_json(self.data_dir / "mhrt_structure.json", structure)
        self._write_json(self.data_dir / "mhrt_metrics.json", metrics)
        return {"structure": structure, "metrics": metrics}

    @staticmethod
    def _clean_name(value):
        cleaned = re.sub(r"\s+", " ", str(value)).strip()
        return NAME_CORRECTIONS.get(cleaned.casefold(), cleaned)

    @staticmethod
    def _active_value(value):
        if isinstance(value, bool):
            return value
        return str(value).strip().casefold() not in {"no", "n", "false", "0", "inactive"}

    @staticmethod
    def _nonnegative_number(value):
        try:
            number = float(value)
        except (TypeError, ValueError):
            return 0
        if number < 0:
            return 0
        return int(number) if number.is_integer() else number
