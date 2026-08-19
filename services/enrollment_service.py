from datetime import datetime
from io import BytesIO
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.timeliness_service import TimelinessService


ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
PROGRAM_COLUMN = "Programs Name"
FIRST_NAME_COLUMN = "Clients First Name"
LAST_NAME_COLUMN = "Clients Last Name"
CLIENT_ID_COLUMN = "Clients Unique Identifier"
START_DATE_COLUMN = "Enrollments Project Start Date"
ENGAGEMENT_DATE_COLUMN = "Entry Screen Date of Engagement"
EXIT_DATE_COLUMN = "Enrollments Project Exit Date"
DESTINATION_COLUMN = "Update/Exit Screen Destination"
DESTINATION_CATEGORY_COLUMN = "Update/Exit Screen Destination Category"
LAST_UPDATED_COLUMN = "Update/Exit Screen Last Updated Date"
CHRONIC_COLUMN = "Entry Screen Chronically Homeless at Project Start - Individual"
FIRST_ENROLLMENT_COLUMN = "Enrollments Is First Enrollment in System (Yes / No)"
ASSIGNED_STAFF_COLUMNS = ["Assigned Staff", "Assigned Staff (* denotes Inactive)"]
PARSED_STAFF_COLUMN = "Full Name"
TEAM_COLUMN = "Team"
SHARED_COLUMN = "Shared Assignment"

EXPECTED_COLUMNS = [
    PROGRAM_COLUMN,
    FIRST_NAME_COLUMN,
    LAST_NAME_COLUMN,
    CLIENT_ID_COLUMN,
    START_DATE_COLUMN,
    ENGAGEMENT_DATE_COLUMN,
    EXIT_DATE_COLUMN,
    DESTINATION_COLUMN,
    DESTINATION_CATEGORY_COLUMN,
    LAST_UPDATED_COLUMN,
    CHRONIC_COLUMN,
    FIRST_ENROLLMENT_COLUMN,
]

TEMPORARY_DESTINATIONS = {
    "Emergency shelter, including hotel or motel paid for with emergency shelter voucher, Host Home shelter",
    "Substance abuse treatment facility or detox center",
}
TEMPORARY_CATEGORY = "Temporary Housing Situations"
PERMANENT_CATEGORY = "Permanent Housing Situations"
NO_EXIT_INTERVIEW = "No exit interview completed"

PIVOTS = [
    (1, "currently_enrolled", "Total Currently Enrolled"),
    (2, "engaged", "Total Engaged"),
    (3, "exits", "Total Exits (All Clients)"),
    (4, "positive_exits", "Positive Exits"),
    (5, "ph_exits", "PH Exits"),
    (6, "no_exit_interview", "No Exit Interview Exits"),
    (7, "chronically_homeless", "Chronically Homeless"),
    (8, "first_enrollment", "First Enrollment in System"),
]

LIGHT_PURPLE_FILL = PatternFill("solid", fgColor="E4D7F5")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FILL = PatternFill("solid", fgColor="B4C6E7")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
THIN_GRAY = Side(style="thin", color="B7B7B7")
TABLE_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


class EnrollmentService:
    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def process(
        self,
        program,
        uploaded_file,
        start_date,
        end_date,
        structure,
        current_metrics,
        team_uploaded_file=None,
        cls_uploaded_file=None,
        services_uploaded_file=None,
        case_notes_raw=None,
        case_notes_clean=None,
        timeliness_pivots=None,
    ):
        program = str(program).strip().upper()
        if program not in {"MHRT", "RRT"}:
            raise ValueError("Program must be MHRT or RRT.")
        start = self._parse_date(start_date, "start")
        end = self._parse_date(end_date, "end")
        if start > end:
            raise ValueError("Start date must be on or before end date.")

        filename = Path(uploaded_file.filename).name
        self._validate_filename_program(program, filename)
        content, extension = self._read_upload(uploaded_file, "Enrollment/Outcomes")
        enrollment_raw = self._load_enrollment_frame(content, extension)
        enrollment_raw = self._clean_headers(enrollment_raw)
        assigned_staff_column = self._validate_columns(enrollment_raw)
        enrollment_raw_for_export = enrollment_raw.copy()
        enrollment_program = self._filter_program(enrollment_raw, program)

        staff_mapping = self._structure_mapping(structure)
        deduplicated = self._deduplicate(enrollment_program)
        expanded = self._expand_assignments(deduplicated, assigned_staff_column, staff_mapping)
        pivots = self._build_pivots(expanded, end)

        metrics = {staff: dict(values) for staff, values in current_metrics.items()}
        roster_staff = [
            staff
            for manager in structure.get("managers", [])
            for staff in manager.get("staff", [])
        ]
        enrollment_fields = [metric_key for _, metric_key, _ in PIVOTS]
        for staff in roster_staff:
            metrics.setdefault(staff, {})
            for field in enrollment_fields:
                metrics[staff][field] = 0

        shared_metrics = {}
        report_staff = []
        for _, metric_key, _ in PIVOTS:
            pivot, shared_pairs = pivots[metric_key]
            for record in pivot.to_dict("records"):
                staff = record[PARSED_STAFF_COLUMN]
                metrics.setdefault(staff, {})
                metrics[staff][metric_key] = int(record["Count"])
                if staff not in report_staff:
                    report_staff.append(staff)
                if (record[TEAM_COLUMN], staff) in shared_pairs:
                    shared_metrics.setdefault(staff, []).append(metric_key)

        cls_raw = self._optional_raw_frame(cls_uploaded_file)
        services_raw = self._optional_raw_frame(services_uploaded_file)
        team_raw = self._optional_raw_frame(team_uploaded_file)
        output_name = self._write_workbook(
            program,
            structure,
            metrics,
            enrollment_raw_for_export,
            deduplicated,
            expanded,
            pivots,
            shared_metrics,
            start,
            end,
            team_raw,
            cls_raw,
            services_raw,
            case_notes_raw,
            case_notes_clean,
            timeliness_pivots,
        )
        return {
            "structure": structure,
            "metrics": metrics,
            "download_name": output_name,
            "source_filename": filename,
            "raw_rows": int(len(enrollment_program)),
            "deduplicated_rows": int(len(deduplicated)),
            "assigned_rows": int(len(expanded)),
            "shared_assignments": int(expanded[SHARED_COLUMN].eq("Yes").sum()) if not expanded.empty else 0,
            "staff_count": int(expanded[PARSED_STAFF_COLUMN].nunique()) if not expanded.empty else 0,
            "report_staff": report_staff,
            "shared_metrics": shared_metrics,
        }

    @staticmethod
    def _clean_text(value):
        if pd.isna(value):
            return ""
        return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()

    @classmethod
    def _normalized(cls, value):
        return cls._clean_text(value).casefold()

    @classmethod
    def _clean_headers(cls, frame):
        frame = frame.copy()
        frame.columns = [cls._clean_text(column) for column in frame.columns]
        return frame

    @staticmethod
    def _parse_date(value, label):
        try:
            return pd.Timestamp(value).normalize()
        except (TypeError, ValueError):
            raise ValueError(f"Select a valid {label} date.") from None

    @staticmethod
    def _validate_filename_program(program, filename):
        normalized = re.sub(r"[^a-z0-9]+", " ", filename.casefold()).strip()
        tokens = set(normalized.split())
        is_mhrt = "mhrt" in tokens or "mobile homelessness resolution" in normalized
        is_rrt = "rrt" in tokens or "rapid response team" in normalized
        if program == "MHRT" and is_rrt and not is_mhrt:
            raise ValueError("This appears to be an RRT Enrollment report. Switch the app to RRT first.")
        if program == "RRT" and is_mhrt and not is_rrt:
            raise ValueError("This appears to be an MHRT Enrollment report. Switch the app to MHRT first.")

    @staticmethod
    def _read_upload(uploaded_file, label):
        extension = Path(uploaded_file.filename).suffix.lower()
        if extension not in ALLOWED_EXTENSIONS:
            raise ValueError(f"{label} upload must be a CSV, XLSX, or XLS file.")
        content = uploaded_file.read()
        if not content:
            raise ValueError(f"The uploaded {label} file is empty.")
        return content, extension

    def _load_enrollment_frame(self, content, extension):
        stream = BytesIO(content)
        if extension == ".csv":
            try:
                return pd.read_csv(stream, dtype=str)
            except UnicodeDecodeError:
                stream.seek(0)
                return pd.read_csv(stream, dtype=str, encoding="latin-1")

        workbook = pd.ExcelFile(stream)
        for sheet_name in workbook.sheet_names:
            preview = pd.read_excel(workbook, sheet_name=sheet_name, header=None, nrows=30)
            for header_row in range(len(preview)):
                values = {self._clean_text(value) for value in preview.iloc[header_row].tolist()}
                if {CLIENT_ID_COLUMN, START_DATE_COLUMN, PROGRAM_COLUMN}.issubset(values):
                    return pd.read_excel(workbook, sheet_name=sheet_name, header=header_row, dtype=str)
        raise ValueError("No Excel sheet contains the expected Enrollment/Outcomes columns.")

    def _validate_columns(self, frame):
        missing = [column for column in EXPECTED_COLUMNS if column not in frame.columns]
        assigned = next((column for column in ASSIGNED_STAFF_COLUMNS if column in frame.columns), None)
        if assigned is None:
            missing.append(ASSIGNED_STAFF_COLUMNS[0])
        if missing:
            raise ValueError("Enrollment file is missing columns: " + ", ".join(missing))
        return assigned

    def _filter_program(self, frame, program):
        frame = frame.copy()
        values = frame[PROGRAM_COLUMN].map(self._normalized)
        if program == "MHRT":
            mask = values.str.contains(r"\bmhrt\b|mobile homelessness resolution", regex=True)
        else:
            mask = values.str.contains(r"\brrt\b|rapid response team", regex=True)
        if frame[PROGRAM_COLUMN].fillna("").astype(str).str.strip().ne("").any():
            frame = frame[mask].copy()
            if frame.empty:
                raise ValueError(f"The Enrollment report has no {program} rows.")
        return frame

    def _structure_mapping(self, structure):
        mapping = {}
        for manager in structure.get("managers", []):
            team = self._clean_text(manager.get("name", "")) or "No Team Assigned"
            for staff in manager.get("staff", []):
                name = self._clean_text(staff)
                if name:
                    mapping[name.casefold()] = (name, team)
        if not mapping:
            raise ValueError("Import or create the staff Team mapping before processing Enrollment/Outcomes.")
        return mapping

    def _deduplicate(self, frame):
        frame = frame.copy()
        for column in [START_DATE_COLUMN, ENGAGEMENT_DATE_COLUMN, EXIT_DATE_COLUMN, LAST_UPDATED_COLUMN]:
            frame[column] = pd.to_datetime(frame[column], errors="coerce")
        frame[CLIENT_ID_COLUMN] = frame[CLIENT_ID_COLUMN].astype("string").str.strip()
        frame = frame[frame[CLIENT_ID_COLUMN].notna() & frame[CLIENT_ID_COLUMN].ne("")].copy()
        destinations = frame[DESTINATION_COLUMN].fillna("").astype(str).str.strip()
        frame.loc[destinations.isin(TEMPORARY_DESTINATIONS), DESTINATION_CATEGORY_COLUMN] = TEMPORARY_CATEGORY
        frame["_active_priority"] = frame[EXIT_DATE_COLUMN].isna()
        frame = frame.sort_values(
            [CLIENT_ID_COLUMN, "_active_priority", START_DATE_COLUMN, EXIT_DATE_COLUMN, LAST_UPDATED_COLUMN],
            ascending=[True, False, False, False, False],
            na_position="last",
            kind="stable",
        )
        return (
            frame.drop_duplicates(subset=[CLIENT_ID_COLUMN], keep="first")
            .drop(columns=["_active_priority"])
            .reset_index(drop=True)
        )

    def _parse_staff(self, value, staff_mapping):
        text = self._clean_text(value)
        if not text:
            return []
        matches = []
        for canonical_name, _ in staff_mapping.values():
            pattern = re.compile(rf"(?<![\w]){re.escape(canonical_name)}(?![\w])", re.IGNORECASE)
            match = pattern.search(text)
            if match:
                matches.append((match.start(), canonical_name))
        matches.sort(key=lambda item: item[0])
        names = list(dict.fromkeys(name for _, name in matches))
        if self._normalized(text).startswith("deactivated per ticket"):
            return names[-1:]
        return names

    def _expand_assignments(self, frame, assigned_staff_column, staff_mapping):
        frame = frame.copy()
        frame["_staff_list"] = frame[assigned_staff_column].apply(
            lambda value: self._parse_staff(value, staff_mapping)
        )
        frame[SHARED_COLUMN] = frame["_staff_list"].map(lambda names: "Yes" if len(names) > 1 else "No")
        frame = frame.explode("_staff_list", ignore_index=True)
        frame = frame[frame["_staff_list"].notna() & frame["_staff_list"].astype(str).str.strip().ne("")].copy()
        frame[PARSED_STAFF_COLUMN] = frame["_staff_list"].astype(str).str.strip()
        frame[TEAM_COLUMN] = frame[PARSED_STAFF_COLUMN].map(
            lambda staff: staff_mapping.get(staff.casefold(), (staff, "No Team Assigned"))[1]
        )
        return frame.drop(columns=["_staff_list"]).reset_index(drop=True)

    def _population(self, expanded, number, end):
        frame = expanded.copy()
        if number == 1:
            mask = pd.to_datetime(frame[EXIT_DATE_COLUMN], errors="coerce").isna()
            mask &= pd.to_datetime(frame[START_DATE_COLUMN], errors="coerce").le(end)
        elif number == 2:
            dates = pd.to_datetime(frame[ENGAGEMENT_DATE_COLUMN], errors="coerce")
            mask = dates.notna() & dates.le(end)
        elif number in {3, 4, 5, 6}:
            dates = pd.to_datetime(frame[EXIT_DATE_COLUMN], errors="coerce")
            mask = dates.notna() & dates.le(end)
            if number == 4:
                allowed = {PERMANENT_CATEGORY.casefold(), TEMPORARY_CATEGORY.casefold()}
                mask &= frame[DESTINATION_CATEGORY_COLUMN].map(self._normalized).isin(allowed)
            elif number == 5:
                mask &= frame[DESTINATION_CATEGORY_COLUMN].map(self._normalized).eq(PERMANENT_CATEGORY.casefold())
            elif number == 6:
                mask &= frame[DESTINATION_COLUMN].map(self._normalized).eq(NO_EXIT_INTERVIEW.casefold())
        elif number == 7:
            mask = frame[CHRONIC_COLUMN].map(self._normalized).eq("yes")
        else:
            mask = frame[FIRST_ENROLLMENT_COLUMN].map(self._normalized).eq("yes")
        return frame[mask].copy()

    def _build_pivots(self, expanded, end):
        results = {}
        for number, metric_key, _ in PIVOTS:
            population = self._population(expanded, number, end)
            pivot = (
                population.groupby([TEAM_COLUMN, PARSED_STAFF_COLUMN], dropna=False)[CLIENT_ID_COLUMN]
                .count()
                .rename("Count")
                .reset_index()
                if not population.empty
                else pd.DataFrame(columns=[TEAM_COLUMN, PARSED_STAFF_COLUMN, "Count"])
            )
            shared = set(
                population.loc[population[SHARED_COLUMN].eq("Yes"), [TEAM_COLUMN, PARSED_STAFF_COLUMN]]
                .drop_duplicates()
                .itertuples(index=False, name=None)
            )
            results[metric_key] = (pivot, shared)
        return results

    def _optional_raw_frame(self, uploaded_file):
        if not uploaded_file or not uploaded_file.filename:
            return None
        content, extension = self._read_upload(uploaded_file, "supporting report")
        stream = BytesIO(content)
        if extension == ".csv":
            try:
                return pd.read_csv(stream, dtype=str)
            except UnicodeDecodeError:
                stream.seek(0)
                return pd.read_csv(stream, dtype=str, encoding="latin-1")
        workbook = pd.ExcelFile(stream)
        best_frame = None
        best_score = -1
        for sheet_name in workbook.sheet_names:
            candidate = pd.read_excel(workbook, sheet_name=sheet_name, dtype=str)
            populated = candidate.dropna(how="all").dropna(axis="columns", how="all")
            nonempty_cells = int(populated.notna().sum().sum())
            score = nonempty_cells + (len(populated) * max(len(populated.columns), 1))
            if score > best_score:
                best_frame = candidate
                best_score = score
        return best_frame if best_frame is not None else pd.DataFrame()

    def _write_workbook(
        self, program, structure, metrics, raw, clean, assigned, pivots, shared_metrics,
        start, end, team_raw, cls_raw, services_raw, case_notes_raw, case_notes_clean,
        timeliness_pivots,
    ):
        program_dir = self.output_dir / program
        program_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{program}_Weekly_Combined_{start:%Y%m%d}_{end:%Y%m%d}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = program_dir / filename
        workbook = Workbook()
        weekly = workbook.active
        weekly.title = "Weekly Table"
        self._write_weekly_table(weekly, program, structure, metrics, shared_metrics, start, end)

        dashboard = workbook.create_sheet("Pivots")
        dashboard["A1"] = f"{program} PIVOT TABLES"
        dashboard["A1"].font = Font(size=14, bold=True)
        dashboard["A2"] = "Measure Start Date"
        dashboard["B2"] = start.to_pydatetime()
        dashboard["B2"].number_format = "mm/dd/yyyy"
        dashboard["A3"] = "Measure End Date"
        dashboard["B3"] = end.to_pydatetime()
        dashboard["B3"].number_format = "mm/dd/yyyy"

        row = 5
        for metric_key, title in [
            ("clients_served", "Total Clients Served"),
            ("cls_assessments", "Instances of Service (CLS Assessments)"),
            ("services", "Total Services Provided"),
            ("referrals_to_shelter", "Referral to Shelter"),
        ]:
            row = self._write_saved_metric_pivot(dashboard, row, 1, title, metric_key, structure, metrics)

        enrollment_row = 5
        for _, metric_key, title in PIVOTS:
            pivot, shared = pivots[metric_key]
            enrollment_row = self._write_pivot(dashboard, enrollment_row, 5, title, pivot, shared)
        dashboard.cell(row=enrollment_row, column=5, value="Shared assignment").fill = LIGHT_PURPLE_FILL

        if timeliness_pivots:
            TimelinessService.write_pivots(
                dashboard,
                timeliness_pivots["case_notes"],
                timeliness_pivots["average_days"],
                start_row=max(row, enrollment_row) + 2,
                start_column=1,
            )

        for column, width in {1: 28, 2: 28, 3: 12, 5: 28, 6: 28, 7: 12}.items():
            dashboard.column_dimensions[get_column_letter(column)].width = width
        dashboard.freeze_panes = "A5"
        dashboard.sheet_view.showGridLines = False

        if team_raw is not None:
            self._write_frame(workbook, "Team_Raw", team_raw)
        if cls_raw is not None:
            self._write_frame(workbook, "CLS_Raw", cls_raw)
        if services_raw is not None:
            self._write_frame(workbook, "Services_Raw", services_raw)
        self._write_frame(workbook, "Enrollment_Raw", raw)
        if case_notes_raw is not None:
            self._write_frame(workbook, "Case_Notes_Raw", case_notes_raw)
        if case_notes_clean is not None:
            self._write_frame(workbook, "Case_Notes_Clean", case_notes_clean)
        self._write_frame(workbook, "Enrollment_Clean", clean)
        self._write_frame(workbook, "Enrollment_Assigned", assigned)
        workbook.save(path)
        return filename

    def _write_weekly_table(self, sheet, program, structure, metrics, shared_metrics, start, end):
        columns = [
            ("clients_served", "Total Clients Served", "15-20", "00B050"),
            ("services", "Total Services Provided", "45-60", "F4B183"),
            ("referrals_to_shelter", "Referral to Shelter", "", "F8CBAD"),
            ("currently_enrolled", "Total Currently Enrolled", "20+", "9DC3E6"),
            ("engaged", "Total Engaged", "15-20", "FFE699"),
            ("exits", "Total Exits (All Clients)", "-", "9DC3E6"),
            ("positive_exits", "Positive Exits", "-", "9DC3E6"),
            ("ph_exits", "PH Exits", "-", "9DC3E6"),
            ("no_exit_interview", "No Exit Interview Exits", "0", "9DC3E6"),
        ]
        if program == "MHRT":
            columns.extend([
                ("chronically_homeless", "Chronically Homeless", "", "C6E0B4"),
                ("first_enrollment", "First Enrollment in System", "-", "C6E0B4"),
            ])
        columns.extend([
            ("__spacer__", "", "", "A6A6A6"),
            ("cls_assessments", "Instances of Service (CLS Assessments)", "45-60", "00B050"),
            ("case_notes", "Case Notes (Enrollment Level)", "45-60", "F4B6D2"),
            ("average_days", "Average # of Days between Service Provided and Case Note Entered", "1-3 days", "F4B6D2"),
        ])

        sheet.cell(1, 1, "Outreach Outcomes")
        sheet.cell(1, 1).font = Font(bold=True)
        sheet.cell(1, 1).alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(2, 1, "CSO Monthly Goals")
        sheet.cell(2, 1).font = Font(bold=True)
        sheet.cell(2, 1).fill = PatternFill("solid", fgColor="FFF200")
        for index, (_, label, goal, color) in enumerate(columns, start=2):
            header = sheet.cell(1, index, label)
            header.fill = PatternFill("solid", fgColor=color)
            header.font = Font(bold=True)
            header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            goal_cell = sheet.cell(2, index, goal)
            goal_cell.fill = PatternFill("solid", fgColor="FFF200" if label else "A6A6A6")
            goal_cell.font = Font(bold=True)
            goal_cell.alignment = Alignment(horizontal="center")

        active = structure.get("active", {})
        row = 3
        all_staff_rows = []
        for manager_index, manager in enumerate(structure.get("managers", [])):
            visible_staff = []
            for staff in manager.get("staff", []):
                values = metrics.get(staff, {})
                has_activity = any(float(values.get(key, 0) or 0) != 0 for key, _, _, _ in columns if key != "__spacer__")
                if active.get(staff, True) or has_activity:
                    visible_staff.append(staff)
            if not visible_staff:
                continue
            manager_staff_rows = []
            for staff in visible_staff:
                sheet.cell(row, 1, staff)
                sheet.cell(row, 1).fill = PatternFill("solid", fgColor="E7E6E6")
                for index, (key, _, _, _) in enumerate(columns, start=2):
                    cell = sheet.cell(row, index)
                    if key == "__spacer__":
                        cell.fill = PatternFill("solid", fgColor="A6A6A6")
                        continue
                    value = metrics.get(staff, {}).get(key, 0)
                    cell.value = "-" if key == "average_days" and value is None else value
                    cell.number_format = "0.00" if key == "average_days" else "0"
                    cell.alignment = Alignment(horizontal="center")
                    if key in shared_metrics.get(staff, []):
                        cell.fill = LIGHT_PURPLE_FILL
                manager_staff_rows.append(row)
                all_staff_rows.append(row)
                row += 1

            sheet.cell(row, 1, f"{manager.get('name', '')} Total")
            sheet.cell(row, 1).font = Font(bold=True)
            for index, (key, _, _, _) in enumerate(columns, start=2):
                cell = sheet.cell(row, index)
                if key == "__spacer__":
                    cell.fill = PatternFill("solid", fgColor="A6A6A6")
                    continue
                references = ",".join(f"{get_column_letter(index)}{staff_row}" for staff_row in manager_staff_rows)
                if key == "average_days":
                    cell.value = f'=IFERROR(AVERAGE({references}),"-")'
                else:
                    cell.value = f"=SUM({references})"
                cell.number_format = "0.00" if key == "average_days" else "0"
                cell.font = Font(bold=True)
            for cell in sheet[row]:
                if cell.column != next((i for i, item in enumerate(columns, start=2) if item[0] == "__spacer__"), -1):
                    cell.fill = PatternFill("solid", fgColor="BFBFBF")
            row += 2

        sheet.cell(row, 1, f"{program} Total")
        sheet.cell(row, 1).font = Font(bold=True)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor="FFF200")
        for index, (key, _, _, _) in enumerate(columns, start=2):
            cell = sheet.cell(row, index)
            if key == "__spacer__":
                cell.fill = PatternFill("solid", fgColor="A6A6A6")
                continue
            references = ",".join(f"{get_column_letter(index)}{staff_row}" for staff_row in all_staff_rows)
            if not references:
                cell.value = "-" if key == "average_days" else 0
            elif key == "average_days":
                cell.value = f'=IFERROR(AVERAGE({references}),"-")'
            else:
                cell.value = f"=SUM({references})"
            cell.number_format = "0.00" if key == "average_days" else "0"
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFF200")
        row += 1
        sheet.cell(row, 1, "Date range:")
        sheet.cell(row, 1).font = Font(bold=True)
        sheet.cell(row, 2, f"{start:%m/%d/%Y}-{end:%m/%d/%Y}")

        last_column = len(columns) + 1
        for row_cells in sheet.iter_rows(min_row=1, max_row=row, min_col=1, max_col=last_column):
            for cell in row_cells:
                cell.border = Border(
                    left=Side(style="thin", color="808080"),
                    right=Side(style="thin", color="808080"),
                    top=Side(style="thin", color="808080"),
                    bottom=Side(style="thin", color="808080"),
                )
        sheet.row_dimensions[1].height = 90
        sheet.column_dimensions["A"].width = 27
        for index, (key, _, _, _) in enumerate(columns, start=2):
            sheet.column_dimensions[get_column_letter(index)].width = 4 if key == "__spacer__" else 14
        sheet.freeze_panes = "B3"
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = "1:2"
        sheet.print_area = f"A1:{get_column_letter(last_column)}{row}"

    def _write_saved_metric_pivot(self, sheet, row, column, title, metric_key, structure, metrics):
        records = []
        for manager in structure.get("managers", []):
            for staff in manager.get("staff", []):
                count = int(float(metrics.get(staff, {}).get(metric_key, 0) or 0))
                if count:
                    records.append({TEAM_COLUMN: manager.get("name", ""), PARSED_STAFF_COLUMN: staff, "Count": count})
        frame = pd.DataFrame(records, columns=[TEAM_COLUMN, PARSED_STAFF_COLUMN, "Count"])
        return self._write_pivot(sheet, row, column, title, frame, set())

    def _write_pivot(self, sheet, row, column, title, frame, shared):
        title_cell = sheet.cell(row=row, column=column, value=title)
        title_cell.fill = SECTION_FILL
        title_cell.font = Font(bold=True)
        for offset, header in enumerate([TEAM_COLUMN, PARSED_STAFF_COLUMN, "Count"]):
            cell = sheet.cell(row=row + 1, column=column + offset, value=header)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
            cell.border = TABLE_BORDER
        current = row + 2
        if frame.empty:
            sheet.cell(row=current, column=column, value="No matching records")
            current += 1
        else:
            frame = frame.sort_values([TEAM_COLUMN, PARSED_STAFF_COLUMN], kind="stable")
            for record in frame.to_dict("records"):
                pair = (record[TEAM_COLUMN], record[PARSED_STAFF_COLUMN])
                for offset, value in enumerate([pair[0], pair[1], int(record["Count"])]):
                    cell = sheet.cell(row=current, column=column + offset, value=value)
                    cell.border = TABLE_BORDER
                    if pair in shared:
                        cell.fill = LIGHT_PURPLE_FILL
                current += 1
            for offset, value in enumerate(["Grand Total", "", int(frame["Count"].sum())]):
                cell = sheet.cell(row=current, column=column + offset, value=value)
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)
                cell.border = TABLE_BORDER
            current += 1
        return current + 2

    @staticmethod
    def _write_frame(workbook, name, frame):
        sheet = workbook.create_sheet(name[:31])
        for column_number, header in enumerate(frame.columns, start=1):
            cell = sheet.cell(1, column_number, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.alignment = Alignment(wrap_text=True)
        for row_number, values in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_number, value in enumerate(values, start=1):
                if pd.isna(value):
                    value = None
                elif isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                cell = sheet.cell(row_number, column_number, value)
                if isinstance(value, datetime):
                    cell.number_format = "yyyy-mm-dd"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
