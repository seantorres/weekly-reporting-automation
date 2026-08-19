from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app import create_app
from services.structure_service import StructureService


def test_structure_creates_report_staff(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path})
    client = app.test_client()
    payload = {
        "managers": [
            {"name": "Manager One", "staff": ["Staff A", "Staff B"]}
        ]
    }

    response = client.put("/api/programs/MHRT/structure", json=payload)
    assert response.status_code == 200
    assert response.get_json()["managers"][0]["staff"] == ["Staff A", "Staff B"]

    saved = client.get("/api/programs/MHRT/structure").get_json()
    assert saved["program"] == "MHRT"
    assert len(saved["managers"]) == 1


def test_programs_are_kept_separate(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path})
    client = app.test_client()
    client.put("/api/programs/MHRT/structure", json={"managers": [{"name": "M", "staff": ["A"]}]})

    rrt = client.get("/api/programs/RRT/structure").get_json()
    assert rrt == {"program": "RRT", "managers": [], "active": {}}


def test_api_responses_are_not_cached(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path})
    response = app.test_client().get("/api/programs/MHRT/metrics")
    assert response.headers["Cache-Control"] == "no-store"


def test_duplicate_staff_are_removed(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path})
    client = app.test_client()
    payload = {"managers": [
        {"name": "M1", "staff": ["Staff A"]},
        {"name": "M2", "staff": ["staff a", "Staff B"]},
    ]}
    saved = client.put("/api/programs/MHRT/structure", json=payload).get_json()
    assert saved["managers"][1]["staff"] == ["Staff B"]


def test_jayne_lee_is_migrated_to_jayna_lee_in_roster_and_metrics(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    structure = client.put(
        "/api/programs/MHRT/structure",
        json={
            "managers": [{"name": "No Team", "staff": ["Jayne Lee"]}],
            "active": {"Jayne Lee": False},
            "service_users": {"Jayne Lee": "jlee"},
        },
    ).get_json()
    client.put("/api/programs/MHRT/metrics", json={"Jayne Lee": {"services": 4}})

    assert structure["managers"][0]["staff"] == ["Jayna Lee"]
    assert structure["active"] == {"Jayna Lee": False}
    assert client.get("/api/programs/MHRT/service-users").get_json() == {"Jayna Lee": "jlee"}
    assert client.get("/api/programs/MHRT/metrics").get_json()["Jayna Lee"]["services"] == 4


def test_mhrt_mock_populates_structure_and_metrics(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path})
    client = app.test_client()

    response = client.post("/api/programs/MHRT/mock")
    assert response.status_code == 200
    result = response.get_json()
    assert len(result["structure"]["managers"]) == 3
    assert len(result["metrics"]) == 12
    assert result["metrics"]["Alex Rivera"]["clients_served"] == 15


def test_rrt_rejects_mhrt_mock(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path})
    response = app.test_client().post("/api/programs/RRT/mock")
    assert response.status_code == 400


def test_structure_preset_round_trip(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path, "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    structure = {
        "program": "MHRT",
        "managers": [{"name": "Manager", "staff": ["Staff A"]}],
        "service_users": {"Staff A": "staffuser"},
    }
    saved = client.post("/api/programs/MHRT/presets", json={"name": "Current", "structure": structure})
    assert saved.status_code == 200
    assert client.get("/api/programs/MHRT/presets").get_json()["presets"] == ["Current"]
    loaded = client.post("/api/programs/MHRT/presets/load", json={"name": "Current"})
    loaded_result = loaded.get_json()
    assert loaded_result["structure"]["managers"][0]["staff"] == ["Staff A"]
    assert loaded_result["service_users"] == {"Staff A": "staffuser"}


def test_cls_csv_populates_two_metrics_and_download(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    csv_data = (
        "Unique ID,Staff Created,Date of Contact,team\n"
        "A1,Staff A,2026-07-01,Manager A\n"
        "A1,Staff A,2026-07-02,Manager A\n"
        "A2,Staff A,2026-07-03,Manager A\n"
        "A2,Staff B,2026-07-02,Manager B\n"
        "B1,Staff B,2026-06-30,Manager B\n"
    ).encode()
    response = client.post(
        "/api/programs/MHRT/cls/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-03",
            "cls_file": (BytesIO(csv_data), "mhrt_cls.csv"),
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    result = response.get_json()
    assert result["metrics"]["Staff A"]["cls_assessments"] == 3
    assert result["metrics"]["Staff A"]["clients_served"] == 2
    assert result["metrics"]["Staff B"]["cls_assessments"] == 1
    # A2 belongs to Staff A because that is the newest row for that client.
    assert result["metrics"]["Staff B"]["clients_served"] == 0
    assert result["unique_clients"] == 2
    assert result["structure"]["managers"] == []
    assert result["unmatched_staff"] == ["Staff A", "Staff B"]
    assert result["download_name"].endswith(".xlsx")
    download = client.get(result["download_url"])
    assert download.status_code == 200
    workbook = load_workbook(BytesIO(download.data), data_only=False)
    assert workbook.sheetnames == ["Unique Summary", "Summary", "Filtered", "Unique"]
    assert workbook["Summary"].max_row == 8
    assert workbook["Filtered"].max_row == 5
    assert workbook["Unique"].max_row == 3
    summary_values = {
        workbook["Summary"].cell(row, 1).value: workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    unique_values = {
        workbook["Unique Summary"].cell(row, 1).value: workbook["Unique Summary"].cell(row, 2).value
        for row in range(1, workbook["Unique Summary"].max_row + 1)
    }
    assert summary_values["Manager A"] == 3
    assert summary_values["Staff B"] == 1
    assert unique_values["Staff A"] == 2
    assert unique_values["Grand Total"] == 2


def test_optional_team_file_builds_manager_structure(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    cls_data = (
        "Unique ID,Assigned Staff (* denotes Inactive),Staff Created,Date of Contact\n"
        "A1,Assigned A,Assigned A,2026-07-01\n"
        "A2,Assigned A,Assigned A,2026-07-02\n"
    ).encode()
    team_data = (
        "Staff Name,Program,Team\n"
        "Assigned A,MHRT,Manager Alpha\n"
    ).encode()
    response = client.post(
        "/api/programs/MHRT/cls/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-02",
            "cls_file": (BytesIO(cls_data), "cls.csv"),
            "team_file": (BytesIO(team_data), "Team.csv"),
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    result = response.get_json()
    assert result["team_file_used"] is True
    assert result["structure"]["managers"] == [{"name": "Manager Alpha", "staff": ["Assigned A"]}]
    assert result["metrics"]["Assigned A"]["cls_assessments"] == 2


def test_team_csv_populates_step_one_without_cls(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    team_data = (
        "Staff Name,Program,Team/Manager\n"
        "Staff A,MHRT,Manager One\n"
        "Staff B,MHRT,Manager One\n"
        "Staff C,MHRT,Manager Two\n"
        "RRT Staff,RRT,RRT Manager\n"
    ).encode()
    response = client.post(
        "/api/programs/MHRT/team/import",
        data={"team_file": (BytesIO(team_data), "Team.csv")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    result = response.get_json()
    assert result["staff_count"] == 3
    assert result["manager_count"] == 2
    assert result["structure"]["managers"] == [
        {"name": "Manager One", "staff": ["Staff A", "Staff B"]},
        {"name": "Manager Two", "staff": ["Staff C"]},
    ]
    assert client.get("/api/programs/MHRT/structure").get_json() == result["structure"]


def test_legacy_team_file_uses_first_and_third_columns(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    team_data = (
        "Staff Name,Manager Short Name,Team/Manager\n"
        "Cory Stapleton,Cory,MHRT - Cory\n"
        "Mick Simpson,Autumn,MHRT -Autumn\n"
        "Jayne Lee,no team,No Team\n"
    ).encode()

    response = client.post(
        "/api/programs/MHRT/team/import",
        data={"team_file": (BytesIO(team_data), "Team.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["structure"]["managers"] == [
        {"name": "MHRT - Cory", "staff": ["Cory Stapleton"]},
        {"name": "MHRT -Autumn", "staff": ["Mick Simpson"]},
        {"name": "No Team", "staff": ["Jayna Lee"]},
    ]
    assert result["service_users"] == {}


def test_combined_team_file_imports_program_member_manager_username_and_active(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    team_data = (
        "Program,Member,Manager,Service Creating,Active\n"
        "MHRT,Active Staff,MHRT - Cory,ActiveUser,Yes\n"
        "MHRT,Inactive Staff,MHRT - Autumn,InactiveUser,No\n"
        "MHRT,No Username,MHRT - Payan,,Yes\n"
        "RRT,RRT Staff,RRT Manager,RRTUser,Yes\n"
    ).encode()

    response = client.post(
        "/api/programs/MHRT/team/import",
        data={"team_file": (BytesIO(team_data), "Combined Team.csv")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["structure"] == {
        "program": "MHRT",
        "managers": [
            {"name": "MHRT - Cory", "staff": ["Active Staff"]},
            {"name": "MHRT - Autumn", "staff": ["Inactive Staff"]},
            {"name": "MHRT - Payan", "staff": ["No Username"]},
        ],
        "active": {"Active Staff": True, "Inactive Staff": False, "No Username": True},
    }
    assert result["service_users"] == {
        "Active Staff": "ActiveUser",
        "Inactive Staff": "InactiveUser",
    }


def test_team_export_writes_combined_excel_format(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    response = client.post(
        "/api/programs/MHRT/team/export",
        json={
            "managers": [{"name": "MHRT - Cory", "staff": ["Active Staff", "Former Staff"]}],
            "service_users": {"Active Staff": "ActiveUser"},
            "active": {"Active Staff": True, "Former Staff": False},
        },
    )

    assert response.status_code == 200
    result = response.get_json()
    workbook_response = client.get(result["download_url"])
    workbook = load_workbook(BytesIO(workbook_response.data), data_only=True)
    rows = list(workbook["Team Mapping"].values)
    assert rows == [
        ("Program", "Member", "Manager", "Services User Creating", "Active"),
        ("MHRT", "Active Staff", "MHRT - Cory", "ActiveUser", "Yes"),
        ("MHRT", "Former Staff", "MHRT - Cory", None, "No"),
    ]

    reimported = client.post(
        "/api/programs/MHRT/team/import",
        data={"team_file": (BytesIO(workbook_response.data), result["download_name"])},
        content_type="multipart/form-data",
    )
    assert reimported.status_code == 200
    assert reimported.get_json()["structure"]["active"] == {
        "Active Staff": True,
        "Former Staff": False,
    }


def test_two_tab_cls_workbook_uses_unique_and_instance_tabs(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    instances = pd.DataFrame([
        {"Unique ID": "A1", "Staff Created": "Staff A", "Date of Contact": "2026-07-01"},
        {"Unique ID": "A1", "Staff Created": "Staff A", "Date of Contact": "2026-07-02"},
        {"Unique ID": "A2", "Staff Created": "Staff A", "Date of Contact": "2026-07-03"},
    ])
    unique = pd.DataFrame([
        {"Unique ID": "A1", "Staff Created": "Staff A", "Date of Contact": "2026-07-02"},
        {"Unique ID": "A2", "Staff Created": "Staff A", "Date of Contact": "2026-07-03"},
    ])
    workbook = BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        instances.to_excel(writer, sheet_name="CLS Assessments", index=False)
        unique.to_excel(writer, sheet_name="Unique CLS", index=False)
    workbook.seek(0)

    response = client.post(
        "/api/programs/MHRT/cls/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-03",
            "cls_file": (workbook, "cls_two_tabs.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    result = response.get_json()
    assert result["metrics"]["Staff A"]["clients_served"] == 2
    assert result["metrics"]["Staff A"]["cls_assessments"] == 3


def test_mhrt_cls_filename_cannot_be_processed_as_rrt(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    csv_data = (
        "Unique ID,Staff Created,Date of Contact\n"
        "A1,MHRT Staff,2026-07-01\n"
    ).encode()

    response = client.post(
        "/api/programs/RRT/cls/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-01",
            "cls_file": (BytesIO(csv_data), "MHRT CLS 7.10.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 400
    assert "MHRT CLS report" in response.get_json()["error"]
    assert client.get("/api/programs/RRT/metrics").get_json() == {}
    assert client.get("/api/programs/MHRT/metrics").get_json() == {}


def test_services_populates_two_columns_from_one_deduplicated_file(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    client.put(
        "/api/programs/MHRT/structure",
        json={"managers": [{"name": "MHRT - Cory", "staff": ["Staff A", "Staff B"]}]},
    )
    client.put(
        "/api/programs/MHRT/metrics",
        json={"Staff A": {"clients_served": 9, "cls_assessments": 12}},
    )
    headers = [
        "Programs Name", "Clients Unique Identifier", "Clients First Name", "Clients Last Name",
        "Services Service Added Date", "Services Start Date Date",
        "Services Service Reporting Period Start Date", "Services Service Reporting Period End Date",
        "Services Service Category", "Services Name", "Services User Creating",
        "Services Type Provided", "Service Items Service Item Name", "Services Services Notes",
    ]
    rows = [
        ["MHRT", "A1", "One", "Client", "2026-07-01", "2026-07-01", "2026-07-01", "2026-07-31", "Basic", "Service", "Staff A", "Provided", "Food", "note 1"],
        ["MHRT", "A1", "One", "Client", "2026-07-02", "2026-07-01", "2026-07-01", "2026-07-31", "Basic", "Service", "Staff A", "Provided", "Food", "note 1"],
        ["MHRT", "A1", "One", "Client", "2026-07-02", "2026-07-01", "2026-07-01", "2026-07-31", "Referral", "Service", "Staff A", "Provided", "Shelter Referral", "note 2"],
        ["MHRT", "B1", "Two", "Client", "2026-07-02", "2026-07-02", "2026-07-01", "2026-07-31", "Referral", "Service", "Staff B", "Provided", "Shelter Referral", "note 3"],
        ["MHRT", "C1", "Blank", "Service", "2026-07-02", "2026-07-02", "2026-07-01", "2026-07-31", "Basic", "", "Staff C", "Provided", "Food", "note blank"],
        ["RRT", "R1", "Other", "Program", "2026-07-02", "2026-07-02", "2026-07-01", "2026-07-31", "Basic", "Service", "RRT Staff", "Provided", "Food", "note 4"],
        ["MHRT", "C1", "Outside", "Range", "2026-06-30", "2026-06-30", "2026-06-01", "2026-06-30", "Basic", "Service", "Staff B", "Provided", "Food", "note 5"],
    ]
    csv_data = (",".join(headers) + "\n" + "\n".join(",".join(row) for row in rows)).encode()

    response = client.post(
        "/api/programs/MHRT/services/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-03",
            "services_file": (BytesIO(csv_data), "MHRT Services 7.10.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["rows_in_date_range"] == 3
    assert result["duplicates_removed"] == 1
    assert result["shelter_referrals"] == 2
    assert result["metrics"]["Staff A"]["services"] == 2
    assert result["metrics"]["Staff A"]["referrals_to_shelter"] == 1
    assert result["metrics"]["Staff A"]["clients_served"] == 9
    assert result["metrics"]["Staff A"]["cls_assessments"] == 12
    assert result["metrics"]["Staff B"]["services"] == 1
    assert result["metrics"]["Staff B"]["referrals_to_shelter"] == 1
    assert "Staff C" not in result["metrics"]
    assert "RRT Staff" not in result["metrics"]

    download = client.get(result["download_url"])
    workbook = load_workbook(BytesIO(download.data), data_only=True)
    assert workbook.sheetnames == ["Summary", "Filtered"]
    assert workbook["Filtered"].max_row == 4
    grand_totals = [
        workbook["Summary"].cell(row, 2).value
        for row in range(1, workbook["Summary"].max_row + 1)
        if workbook["Summary"].cell(row, 1).value == "Grand Total"
    ]
    assert grand_totals == [3, 2]


def test_services_excel_finds_headers_below_looker_title_rows(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    row = {
        "Programs Name": "MHRT",
        "Clients Unique Identifier": "A1",
        "Clients First Name": "Test",
        "Clients Last Name": "Client",
        "Services Service Added Date": "2026-07-01",
        "Services Start Date Date": "2026-07-01",
        "Services Service Reporting Period Start Date": "2026-07-01",
        "Services Service Reporting Period End Date": "2026-07-31",
        "Services Service Category": "Referral",
        "Services Name": "Outreach Service",
        "Services User Creating": "Staff A",
        "Services Type Provided": "Provided",
        "Service Items Service Item Name": "Shelter Referral",
        "Services Services Notes": "Referral completed",
    }
    workbook = BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame([["Outreach Reporting - Services and Geolocations"]]).to_excel(
            writer, sheet_name="Report", index=False, header=False,
        )
        pd.DataFrame([row]).to_excel(writer, sheet_name="Report", index=False, startrow=4)
    workbook.seek(0)

    response = client.post(
        "/api/programs/MHRT/services/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-10",
            "services_file": (workbook, "Outreach Reporting - Services and Geolocations.xlsx"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["rows_in_date_range"] == 1
    assert result["shelter_referrals"] == 1


def test_manual_service_user_mapping_persists_by_program(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    response = client.put(
        "/api/programs/RRT/structure",
        json={
            "managers": [{"name": "RRT", "staff": ["RRT Staff"]}],
            "service_users": {"RRT Staff": "rrtuser"},
        },
    )
    assert response.status_code == 200
    assert client.get("/api/programs/RRT/service-users").get_json() == {"RRT Staff": "rrtuser"}
    assert client.get("/api/programs/MHRT/service-users").get_json() == {}


def test_default_service_users_replace_old_self_references(tmp_path):
    mapping_dir = tmp_path / "sample_data"
    mapping_dir.mkdir()
    (mapping_dir / "MHRT_service_user_mapping.csv").write_text(
        "Services User Creating,Fullname,Team\n"
        "MCano,Maria Cano,MHRT -Autumn\n"
        "CStapleton,Cory Stapleton,MHRT - Cory\n",
        encoding="utf-8",
    )
    service = StructureService(tmp_path / "data", mapping_dir)
    service.save_service_users(
        "MHRT",
        {"MCano": "MCano", "Cory Stapleton": "CStapleton"},
    )

    assert service.load_service_users("MHRT") == {
        "Maria Cano": "MCano",
        "Cory Stapleton": "CStapleton",
    }


def test_username_fullname_team_file_populates_services_lookup(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    team_data = (
        "Services User Creating,Fullname,Team\n"
        "CStapleton,Cory Stapleton,MHRT - Cory\n"
        "LMajor,Larry Major,MHRT -Payan\n"
        "LMajor,Larry Major,MHRT - Aisha\n"
    ).encode()
    imported = client.post(
        "/api/programs/MHRT/team/import",
        data={"team_file": (BytesIO(team_data), "MHRT_service_user_mapping.csv")},
        content_type="multipart/form-data",
    )
    assert imported.status_code == 200
    result = imported.get_json()
    assert result["structure"]["managers"] == [
        {"name": "MHRT - Cory", "staff": ["Cory Stapleton"]},
        {"name": "MHRT -Payan", "staff": ["Larry Major"]},
    ]
    assert result["service_users"] == {
        "Cory Stapleton": "CStapleton",
        "Larry Major": "LMajor",
    }


def test_services_uses_saved_username_lookup_and_reports_unmatched(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    client.put(
        "/api/programs/MHRT/structure",
        json={
            "managers": [{"name": "MHRT - Cory", "staff": ["Cory Stapleton"]}],
            "service_users": {"Cory Stapleton": "CStapleton"},
        },
    )
    headers = [
        "Programs Name", "Clients Unique Identifier", "Clients First Name", "Clients Last Name",
        "Services Service Added Date", "Services Start Date Date",
        "Services Service Reporting Period Start Date", "Services Service Reporting Period End Date",
        "Services Service Category", "Services Name", "Services User Creating",
        "Services Type Provided", "Service Items Service Item Name", "Services Services Notes",
    ]
    rows = [
        ["MHRT", "A1", "One", "Client", "2026-07-01", "2026-07-01", "2026-07-01", "2026-07-31", "Basic", "Service", "CStapleton", "Provided", "Food", "note 1"],
        ["MHRT", "A2", "Two", "Client", "2026-07-02", "2026-07-02", "2026-07-01", "2026-07-31", "Referral", "Service", "UnknownUser", "Provided", "Shelter Referral", "note 2"],
    ]
    csv_data = (",".join(headers) + "\n" + "\n".join(",".join(row) for row in rows)).encode()
    response = client.post(
        "/api/programs/MHRT/services/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-03",
            "services_file": (BytesIO(csv_data), "MHRT Services.csv"),
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    result = response.get_json()
    assert result["metrics"]["Cory Stapleton"]["services"] == 1
    assert result["metrics"]["UnknownUser"]["referrals_to_shelter"] == 1
    assert result["unmatched_service_users"] == ["UnknownUser"]
    assert result["service_users"]["Cory Stapleton"] == "CStapleton"
    assert result["service_users"]["UnknownUser"] == "UnknownUser"


def test_services_mapping_updates_metrics_without_moving_page_teams(tmp_path):
    app = create_app({
        "TESTING": True,
        "LOAD_DEFAULT_SERVICE_USER_MAPPINGS": True,
        "DEFAULT_SERVICE_USER_MAPPING_DIR": Path(__file__).resolve().parents[1] / "sample_data",
        "DATA_DIR": tmp_path / "data",
        "OUTPUT_DIR": tmp_path / "outputs",
    })
    client = app.test_client()
    client.put(
        "/api/programs/MHRT/structure",
        json={"managers": [
            {"name": "MHRT - Cory", "staff": ["Yanel Morelos", "Terri Sickels", "Autumn McCann"]},
            {"name": "MHRT - Payan", "staff": []},
            {"name": "MHRT - Autumn", "staff": []},
        ], "active": {"Yanel Morelos": True, "Terri Sickels": True, "Autumn McCann": False}},
    )
    headers = [
        "Programs Name", "Clients Unique Identifier", "Clients First Name", "Clients Last Name",
        "Services Service Added Date", "Services Start Date Date",
        "Services Service Reporting Period Start Date", "Services Service Reporting Period End Date",
        "Services Service Category", "Services Name", "Services User Creating",
        "Services Type Provided", "Service Items Service Item Name", "Services Services Notes",
    ]
    rows = [
        ["MHRT", "A1", "One", "Client", "2026-07-01", "2026-07-01", "2026-07-01", "2026-07-31", "Basic", "Service", "YMorelos", "Provided", "Food", "note 1"],
        ["MHRT", "A2", "Two", "Client", "2026-07-01", "2026-07-01", "2026-07-01", "2026-07-31", "Basic", "Service", "TSickels", "Provided", "Food", "note 2"],
        ["MHRT", "A3", "Three", "Client", "2026-07-01", "2026-07-01", "2026-07-01", "2026-07-31", "Basic", "Service", "ALowe", "Provided", "Food", "note 3"],
    ]
    csv_data = (",".join(headers) + "\n" + "\n".join(",".join(row) for row in rows)).encode()

    response = client.post(
        "/api/programs/MHRT/services/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-02",
            "services_file": (BytesIO(csv_data), "MHRT Services.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    managers = {manager["name"]: manager["staff"] for manager in result["structure"]["managers"]}
    assert managers["MHRT - Cory"] == ["Yanel Morelos", "Terri Sickels", "Autumn McCann"]
    assert managers["MHRT - Payan"] == []
    assert managers["MHRT - Autumn"] == []
    assert result["metrics"]["Yanel Morelos"]["services"] == 1
    assert result["metrics"]["Terri Sickels"]["services"] == 1
    assert result["metrics"]["Autumn McCann"]["services"] == 1
    assert result["structure"]["active"]["Autumn McCann"] is False
    assert "Autumn McCann" in result["report_staff"]
    assert result["unmatched_service_users"] == []


def test_enrollment_outcomes_updates_eight_metrics_and_shared_staff(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    client.put(
        "/api/programs/MHRT/structure",
        json={
            "managers": [{
                "name": "MHRT - Autumn",
                "staff": ["Mick Simpson", "Maria Cano", "Lindsey Voeller"],
            }],
            "active": {"Mick Simpson": True, "Maria Cano": True, "Lindsey Voeller": False},
        },
    )
    client.put(
        "/api/programs/MHRT/metrics",
        json={"Mick Simpson": {"clients_served": 7, "services": 11}},
    )
    headers = [
        "Programs Name", "Clients First Name", "Clients Last Name", "Clients Unique Identifier",
        "Enrollments Project Start Date", "Entry Screen Date of Engagement",
        "Enrollments Project Exit Date", "Update/Exit Screen Destination",
        "Update/Exit Screen Destination Category", "Update/Exit Screen Last Updated Date",
        "Entry Screen Chronically Homeless at Project Start - Individual",
        "Enrollments Is First Enrollment in System (Yes / No)", "Assigned Staff",
    ]
    rows = [
        ["MHRT", "One", "Client", "A1", "2026-05-01", "2026-05-02", "2026-06-01", "Other", "Other", "2026-06-02", "No", "No", "Mick Simpson"],
        ["MHRT", "One", "Client", "A1", "2026-06-10", "2026-06-11", "", "", "", "2026-07-01", "Yes", "Yes", "Mick Simpson"],
        ["MHRT", "Two", "Client", "B1", "2026-05-10", "2026-05-11", "2026-06-15", "Emergency shelter, including hotel or motel paid for with emergency shelter voucher, Host Home shelter", "Other", "2026-06-16", "No", "No", "Maria Cano, Mick Simpson"],
        ["MHRT", "Three", "Client", "C1", "2026-05-15", "2026-05-16", "2026-06-20", "Rental by client", "Permanent Housing Situations", "2026-06-21", "No", "No", "Deactivated per ticket 75705 on 2-3-2026 Lindsey Voeller, Mick Simpson.1"],
        ["MHRT", "Four", "Client", "D1", "2026-05-20", "2026-05-21", "2026-06-25", "No exit interview completed", "Other", "2026-06-26", "No", "No", "Maria Cano"],
        ["RRT", "Other", "Program", "R1", "2026-05-01", "2026-05-01", "", "", "", "2026-05-01", "Yes", "Yes", "RRT Staff"],
    ]
    csv_data = (",".join(headers) + "\n" + "\n".join(
        ",".join(f'"{value}"' if "," in value else value for value in row) for row in rows
    )).encode()

    response = client.post(
        "/api/programs/MHRT/enrollment/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-10",
            "enrollment_file": (BytesIO(csv_data), "MHRT Enrollment Outcomes.csv"),
            "team_file": (
                BytesIO(
                    b"Program,Member,Manager,Services User Creating,Active\n"
                    b"MHRT,Mick Simpson,MHRT - Autumn,MSimpson,Yes\n"
                ),
                "Team.csv",
            ),
            "cls_file": (BytesIO(b"Raw CLS\nvalue\n"), "CLS.csv"),
            "services_file": (BytesIO(b"Raw Services\nvalue\n"), "Services.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    mick = result["metrics"]["Mick Simpson"]
    maria = result["metrics"]["Maria Cano"]
    assert mick["clients_served"] == 7
    assert mick["services"] == 11
    assert mick["currently_enrolled"] == 1
    assert mick["engaged"] == 3
    assert mick["exits"] == 2
    assert mick["positive_exits"] == 2
    assert mick["ph_exits"] == 1
    assert mick["no_exit_interview"] == 0
    assert mick["chronically_homeless"] == 1
    assert mick["first_enrollment"] == 1
    assert maria["engaged"] == 2
    assert maria["exits"] == 2
    assert maria["positive_exits"] == 1
    assert maria["no_exit_interview"] == 1
    assert result["deduplicated_rows"] == 4
    assert result["assigned_rows"] == 5
    assert result["shared_assignments"] == 2
    assert "Lindsey Voeller" not in result["report_staff"]
    assert set(result["shared_metrics"]["Mick Simpson"]) >= {"engaged", "exits", "positive_exits"}
    assert set(result["shared_metrics"]["Maria Cano"]) >= {"engaged", "exits", "positive_exits"}

    download = client.get(result["download_url"])
    workbook = load_workbook(BytesIO(download.data), data_only=True)
    assert workbook.sheetnames == [
        "Weekly Table", "Pivots", "Team_Raw", "CLS_Raw", "Services_Raw",
        "Enrollment_Raw", "Enrollment_Clean", "Enrollment_Assigned",
    ]
    assert workbook["Enrollment_Raw"].max_row == 7
    assert workbook["Team_Raw"]["A1"].value == "Program"
    weekly = workbook["Weekly Table"]
    assert weekly["A1"].value == "Outreach Outcomes"
    assert weekly["A2"].value == "CSO Monthly Goals"
    assert weekly.page_setup.orientation == "landscape"
    assert weekly.page_setup.fitToWidth == 1
    assert weekly.print_area
    assert any(cell.value == "Mick Simpson" for row in weekly.iter_rows() for cell in row)
    purple_cells = [
        cell
        for row in workbook["Pivots"].iter_rows()
        for cell in row
        if cell.fill.fgColor.rgb in {"00E4D7F5", "E4D7F5"}
    ]
    assert purple_cells


def test_enrollment_rejects_file_for_other_program(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    response = client.post(
        "/api/programs/RRT/enrollment/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-10",
            "enrollment_file": (BytesIO(b"anything"), "MHRT Enrollment.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert response.status_code == 400
    assert "MHRT Enrollment report" in response.get_json()["error"]


def test_timeliness_builds_team_pivots_and_blanks_negative_tim(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    structure = {
        "program": "MHRT",
        "managers": [{"name": "MHRT - Cory", "staff": ["Cory Stapleton"]}],
        "active": {"Cory Stapleton": True},
    }
    assert client.put("/api/programs/MHRT/structure", json=structure).status_code == 200
    headers = [
        "Clients Unique Identifier", "Clients Last Name", "Programs Name",
        "Enrollments Project Start Date", "Enrollments Project Exit Date",
        "Client Notes - Enrollment Level Case Note Date",
        "Client Notes - Enrollment Level Date Added Date",
        "Client Notes - Enrollment Level Staff Full Name",
        "Client Notes - Enrollment Level Note",
    ]
    rows = [
        ["A1", "One", "MHRT", "2026-06-01", "", "2026-07-02", "2026-07-04", "Cory Stapleton", "First note"],
        ["A1", "One", "MHRT", "2026-06-01", "", "2026-07-02", "2026-07-04", "Cory Stapleton", "First note"],
        ["A2", "Two", "MHRT", "2026-06-02", "", "2026-07-03", "2026-07-02", "Cory Stapleton", "Negative date"],
        ["R1", "Other", "RRT", "2026-06-03", "", "2026-07-04", "2026-07-05", "RRT Staff", "Other program"],
    ]
    csv_data = (",".join(headers) + "\n" + "\n".join(
        ",".join(f'"{value}"' for value in row) for row in rows
    )).encode()

    response = client.post(
        "/api/programs/MHRT/timeliness/process",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-10",
            "case_notes_file": (BytesIO(csv_data), "MHRT Case Notes.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["rows_in_date_range"] == 3
    assert result["deduplicated_rows"] == 2
    assert result["duplicates_removed"] == 1
    assert result["negative_tim_rows"] == 1
    assert result["metrics"]["Cory Stapleton"]["case_notes"] == 2
    assert result["metrics"]["Cory Stapleton"]["average_days"] == 2.0
    assert result["structure"]["managers"] == structure["managers"]

    download = client.get(result["download_url"])
    workbook = load_workbook(BytesIO(download.data), data_only=True)
    assert workbook.sheetnames == ["Pivots", "Case_Notes_Raw", "Case_Notes_Clean"]
    assert workbook["Pivots"]["A1"].value == "Case Notes (Enrollment Level)"
    assert workbook["Pivots"]["C2"].value == "Count of Clients Unique Identifier"
    assert workbook["Pivots"]["C3"].value == 2
    clean = workbook["Case_Notes_Clean"]
    headers_by_column = {cell.value: cell.column for cell in clean[1]}
    tim_values = [
        clean.cell(row, headers_by_column["TIM"]).value
        for row in range(2, clean.max_row + 1)
    ]
    assert sorted(value for value in tim_values if value is not None) == [2]
    assert tim_values.count(None) == 1


def test_timeliness_step_creates_final_combined_workbook(tmp_path):
    app = create_app({"TESTING": True, "DATA_DIR": tmp_path / "data", "OUTPUT_DIR": tmp_path / "outputs"})
    client = app.test_client()
    structure = {
        "program": "MHRT",
        "managers": [{"name": "MHRT - Autumn", "staff": ["Mick Simpson", "Maria Cano"]}],
        "active": {"Mick Simpson": True, "Maria Cano": True},
    }
    assert client.put("/api/programs/MHRT/structure", json=structure).status_code == 200

    case_headers = [
        "Clients Unique Identifier", "Clients Last Name", "Programs Name",
        "Enrollments Project Start Date", "Enrollments Project Exit Date",
        "Client Notes - Enrollment Level Case Note Date",
        "Client Notes - Enrollment Level Date Added Date",
        "Client Notes - Enrollment Level Staff Full Name",
        "Client Notes - Enrollment Level Note",
    ]
    case_row = ["A1", "One", "MHRT", "2026-06-01", "", "2026-07-02", "2026-07-05", "Mick Simpson", "Case note"]
    case_csv = (",".join(case_headers) + "\n" + ",".join(f'"{value}"' for value in case_row) + "\n").encode()

    enrollment_headers = [
        "Programs Name", "Clients First Name", "Clients Last Name", "Clients Unique Identifier",
        "Enrollments Project Start Date", "Entry Screen Date of Engagement",
        "Enrollments Project Exit Date", "Update/Exit Screen Destination",
        "Update/Exit Screen Destination Category", "Update/Exit Screen Last Updated Date",
        "Entry Screen Chronically Homeless at Project Start - Individual",
        "Enrollments Is First Enrollment in System (Yes / No)", "Assigned Staff",
    ]
    enrollment_row = [
        "MHRT", "One", "Client", "A1", "2026-06-01", "2026-06-02", "", "", "",
        "2026-07-01", "Yes", "Yes", "Mick Simpson",
    ]
    enrollment_csv = (
        ",".join(enrollment_headers) + "\n"
        + ",".join(f'"{value}"' for value in enrollment_row) + "\n"
    ).encode()

    cls_workbook = Workbook()
    cls_workbook.active.title = "Blank Summary"
    cls_data = cls_workbook.create_sheet("Filtered")
    cls_data.append(["Staff", "Value"])
    cls_data.append(["Mick Simpson", 1])
    cls_bytes = BytesIO()
    cls_workbook.save(cls_bytes)
    cls_bytes.seek(0)

    response = client.post(
        "/api/programs/MHRT/finalize",
        data={
            "start_date": "2026-07-01",
            "end_date": "2026-07-10",
            "case_notes_file": (BytesIO(case_csv), "MHRT Case Notes.csv"),
            "enrollment_file": (BytesIO(enrollment_csv), "MHRT Enrollment.csv"),
            "team_file": (
                BytesIO(b"Program,Member,Manager,Services User Creating,Active\nMHRT,Mick Simpson,MHRT - Autumn,MSimpson,Yes\n"),
                "Team.csv",
            ),
            "cls_file": (cls_bytes, "MHRT_CLS_Pivots.xlsx"),
            "services_file": (BytesIO(b"Staff,Value\nMick Simpson,1\n"), "Services.csv"),
        },
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    result = response.get_json()
    assert result["metrics"]["Mick Simpson"]["case_notes"] == 1
    assert result["metrics"]["Mick Simpson"]["average_days"] == 3.0
    assert result["metrics"]["Maria Cano"]["case_notes"] == 0
    assert result["metrics"]["Maria Cano"]["average_days"] is None
    download = client.get(result["download_url"])
    workbook = load_workbook(BytesIO(download.data), data_only=True)
    assert workbook.sheetnames == [
        "Weekly Table", "Pivots", "Team_Raw", "CLS_Raw", "Services_Raw",
        "Enrollment_Raw", "Case_Notes_Raw", "Case_Notes_Clean",
        "Enrollment_Clean", "Enrollment_Assigned",
    ]
    assert workbook["Weekly Table"]["O3"].value == 1
    assert workbook["Weekly Table"]["P3"].value == 3.0
    assert workbook["Weekly Table"]["O4"].value == 0
    assert workbook["Weekly Table"]["P4"].value == "-"
    assert workbook["CLS_Raw"]["A1"].value == "Staff"
    assert workbook["CLS_Raw"]["A2"].value == "Mick Simpson"
    pivot_values = [cell.value for row in workbook["Pivots"].iter_rows() for cell in row]
    assert "Count of Clients Unique Identifier" in pivot_values
    assert "Average of TIM" in pivot_values


def test_timeliness_frontend_clears_old_rows_before_rendering_results():
    javascript = (Path(__file__).parents[1] / "static" / "js" / "app.js").read_text(encoding="utf-8")
    template = (Path(__file__).parents[1] / "templates" / "index.html").read_text(encoding="utf-8")
    function_text = javascript.split("async function processTimeliness()", 1)[1].split("function addDownload", 1)[0]
    clear_position = function_text.index('document.querySelector("#report-body").innerHTML = "";')
    metrics_position = function_text.index("currentMetrics = result.metrics || {};")
    render_position = function_text.index("renderTable();")
    assert clear_position < metrics_position < render_position
    assert 'input.value = nullable && (value === null || value === undefined) ? "-"' in javascript
    assert 'await autoProcessWhenReady("cls", processCls' in javascript
    assert 'await autoProcessWhenReady("services", processServices' in javascript
    assert 'await autoProcessWhenReady("enrollment", processEnrollment' in javascript
    assert 'await autoProcessWhenReady("timeliness", processTimeliness' in javascript
    assert 'id="create-complete-excel"' in template
    assert 'id="process-cls"' not in template
